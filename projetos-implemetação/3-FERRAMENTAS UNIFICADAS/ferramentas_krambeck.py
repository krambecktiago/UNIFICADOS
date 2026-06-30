"""
Ferramentas Krambeck — Aplicação Unificada
Módulos:
  • Conferência de Duplicatas (Retorno Bancário vs ERP)
  • Contas a Pagar — Gestão Diária
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import re
import os
import json
import unicodedata
import urllib.request
import urllib.error
import csv
import glob
from datetime import datetime

try:
    from PIL import Image, ImageTk, ImageEnhance
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

# ════════════════════════════════════════════════════════════════════════════════
#  TEMA GLOBAL
# ════════════════════════════════════════════════════════════════════════════════

CORES = {
    "bg":         "#f0f2f5",
    "panel":      "#ffffff",
    "header":     "#ffffff",
    "accent":     "#2563eb",
    "accent_h":   "#1d4ed8",
    "accent2":    "#0ea5e9",
    "text":       "#1e293b",
    "muted":      "#64748b",
    "neutro":     "#f59e0b",
    "entry_bg":   "#f8fafc",
    "entry_fg":   "#1e293b",
    "sidebar":    "#1e293b",
    "success":    "#16a34a",
    "success_bg": "#dcfce7",
    "danger":     "#dc2626",
    "danger_bg":  "#fee2e2",
    "border":     "#e2e8f0",
}


# ════════════════════════════════════════════════════════════════════════════════
#  BOTÃO ARREDONDADO (Canvas)
# ════════════════════════════════════════════════════════════════════════════════

class BotaoRound(tk.Canvas):
    """Botão moderno com bordas arredondadas via Canvas."""
    def __init__(self, parent, text, command,
                 largura=200, altura=38, raio=10,
                 cor=None, cor_hover=None,
                 cor_texto="#ffffff",
                 fonte=("Segoe UI", 10, "bold"), **kwargs):
        cor       = cor       or CORES["accent"]
        cor_hover = cor_hover or CORES["accent_h"]
        try:
            bg_pai = parent.cget("bg")
        except Exception:
            bg_pai = CORES["bg"]
        super().__init__(parent, width=largura, height=altura,
                         bg=bg_pai, highlightthickness=0, **kwargs)
        self._cmd   = command
        self._cor   = cor
        self._corh  = cor_hover
        self._txt   = text
        self._fonte = fonte
        self._ctxt  = cor_texto
        self._l     = largura
        self._h     = altura
        self._r     = raio
        self._desenhar(False)
        self.bind("<Enter>",    lambda e: self._desenhar(True))
        self.bind("<Leave>",    lambda e: self._desenhar(False))
        self.bind("<Button-1>", lambda e: command())
        self.config(cursor="hand2")

    def _rr(self, x1, y1, x2, y2, r, cor):
        self.create_arc(x1,     y1,     x1+2*r, y1+2*r, start=90,  extent=90, fill=cor, outline=cor)
        self.create_arc(x2-2*r, y1,     x2,     y1+2*r, start=0,   extent=90, fill=cor, outline=cor)
        self.create_arc(x1,     y2-2*r, x1+2*r, y2,     start=180, extent=90, fill=cor, outline=cor)
        self.create_arc(x2-2*r, y2-2*r, x2,     y2,     start=270, extent=90, fill=cor, outline=cor)
        self.create_rectangle(x1+r, y1,   x2-r, y2,   fill=cor, outline=cor)
        self.create_rectangle(x1,   y1+r, x2,   y2-r, fill=cor, outline=cor)

    def _desenhar(self, hover):
        self.delete("all")
        cor = self._corh if hover else self._cor
        self._rr(1, 1, self._l-1, self._h-1, self._r, cor)
        self.create_text(self._l//2, self._h//2,
                         text=self._txt, fill=self._ctxt, font=self._fonte)

    def config_text(self, text):
        self._txt = text
        self._desenhar(False)


# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 1 — CONFERÊNCIA DE DUPLICATAS
# ════════════════════════════════════════════════════════════════════════════════

def _norm_dup(num):
    """Remove zeros à esquerda mantendo prefixo R."""
    num = str(num).strip()
    return re.sub(r'^0+', '', num) or '0'

def _ler_excel_banco(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    dados = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        if row[0] and row[0] != 'Seu Número':
            norm = _norm_dup(row[0])
            dados[norm] = {
                'duplicata':     row[0],
                'pagador':       str(row[7]) if row[7] else '',
                'valor_titulo':  str(row[4]) if row[4] else '',
                'valor_cobrado': str(row[5]) if row[5] else '',
                'vencimento':    str(row[2]) if row[2] else '',
                'liquidacao':    str(row[3]) if row[3] else '',
            }
    return dados

def _ler_txt_erp(path):
    dados = {}
    pat = re.compile(r'\b(\d{7,}\.\d+|[R]\d+\.\d+|\d+/\d+\.\d+)\b')
    with open(path, encoding='latin-1') as f:
        for line in f:
            m = pat.search(line)
            if m:
                dup = m.group(1)
                dados[_norm_dup(dup)] = dup
    return dados

def _comparar_duplicatas(banco, erp):
    return [info for norm, info in banco.items() if norm not in erp]


class FrameConferirDuplicatas(tk.Frame):

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=CORES["bg"], **kwargs)
        self._build()

    def _build(self):
        tk.Label(self,
                 text="Selecione o retorno bancário (.xlsx) e o fluxo de caixa ERP (.txt).",
                 bg=CORES["bg"], fg=CORES["muted"], font=("Segoe UI", 9)
                 ).pack(padx=20, pady=(14, 6), anchor="w")

        # ── Seleção de arquivos ──
        frm_f = tk.Frame(self, bg=CORES["panel"], padx=14, pady=12,
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_f.pack(fill="x", padx=20, pady=(0, 8))

        self._var_xlsx = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_txt  = tk.StringVar(value="Nenhum arquivo selecionado")

        self._file_row(frm_f, "Retorno do Banco (.xlsx):",  self._var_xlsx, self._sel_xlsx, 0)
        self._file_row(frm_f, "Fluxo de Caixa ERP (.txt):", self._var_txt,  self._sel_txt,  1)

        # ── Botões ──
        frm_btns = tk.Frame(self, bg=CORES["bg"])
        frm_btns.pack(fill="x", padx=20, pady=(0, 8))

        BotaoRound(frm_btns, "▶  Processar Comparação", self._processar,
                   largura=560, altura=40, raio=12
                   ).pack(side="left", expand=True)

        BotaoRound(frm_btns, "↺  Limpar", self._resetar,
                   largura=130, altura=40, raio=12,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"]
                   ).pack(side="left", padx=(8, 0))

        # ── Treeview resultado ──
        frm_r = tk.Frame(self, bg=CORES["panel"],
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_r.pack(fill="both", expand=True, padx=20, pady=(0, 6))

        style = ttk.Style()
        style.configure("Dup.Treeview",
                        background=CORES["entry_bg"], foreground=CORES["text"],
                        fieldbackground=CORES["entry_bg"],
                        font=("Segoe UI", 10), rowheight=26)
        style.configure("Dup.Treeview.Heading",
                        background=CORES["header"], foreground=CORES["accent2"],
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Dup.Treeview",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "#ffffff")])

        self.tree = ttk.Treeview(frm_r, columns=("status", "dup", "pag", "vtitulo", "vcobrado", "vjuros"),
                                  show="headings", style="Dup.Treeview", height=14)
        self.tree.heading("status",   text="Status",     anchor="center")
        self.tree.heading("dup",      text="Duplicata",  anchor="w")
        self.tree.heading("pag",      text="Pagador",    anchor="w")
        self.tree.heading("vtitulo",  text="Vl. Título", anchor="e")
        self.tree.heading("vcobrado", text="Vl. Pago",   anchor="e")
        self.tree.heading("vjuros",   text="Juros",      anchor="e")
        self.tree.column("status",   width=115, anchor="center", stretch=tk.NO)
        self.tree.column("dup",      width=140, anchor="w",      stretch=tk.NO)
        self.tree.column("pag",      width=250, anchor="w",      stretch=tk.YES)
        self.tree.column("vtitulo",  width=100, anchor="e",      stretch=tk.NO)
        self.tree.column("vcobrado", width=100, anchor="e",      stretch=tk.NO)
        self.tree.column("vjuros",   width=90,  anchor="e",      stretch=tk.NO)

        sb = ttk.Scrollbar(frm_r, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree.tag_configure("ok",  background=CORES["success_bg"], foreground=CORES["success"])
        self.tree.tag_configure("nok", background=CORES["danger_bg"],  foreground=CORES["danger"])

        # ── Status ──
        self._status = tk.StringVar(value="Selecione os arquivos e clique em Processar.")
        tk.Label(self, textvariable=self._status,
                 font=("Segoe UI", 9), bg=CORES["bg"], fg=CORES["muted"], anchor="w"
                 ).pack(fill="x", padx=22, pady=(0, 10))

    def _file_row(self, parent, lbl, var, cmd, row):
        tk.Label(parent, text=lbl, font=("Segoe UI", 9),
                 bg=CORES["panel"], fg=CORES["text"], width=30, anchor="w"
                 ).grid(row=row, column=0, pady=6, sticky="w")
        tk.Label(parent, textvariable=var, font=("Segoe UI", 9),
                 fg=CORES["muted"], bg=CORES["panel"], width=34, anchor="w"
                 ).grid(row=row, column=1, padx=8)
        BotaoRound(parent, "Selecionar…", cmd,
                   largura=110, altura=32, raio=8,
                   cor=CORES["header"], cor_hover=CORES["accent"],
                   cor_texto=CORES["text"],
                   fonte=("Segoe UI", 9)
                   ).grid(row=row, column=2, pady=4)

    def _sel_xlsx(self):
        f = filedialog.askopenfilename(title="Retorno do Banco",
                                       filetypes=[("Excel", "*.xlsx *.xls")])
        if f:
            self._xlsx_path = f
            self._var_xlsx.set(os.path.basename(f))

    def _sel_txt(self):
        f = filedialog.askopenfilename(title="Fluxo de Caixa ERP",
                                       filetypes=[("Texto", "*.txt")])
        if f:
            self._txt_path = f
            self._var_txt.set(os.path.basename(f))

    def _processar(self):
        xp = getattr(self, '_xlsx_path', None)
        tp = getattr(self, '_txt_path',  None)
        if not xp or not tp:
            messagebox.showwarning("Atenção", "Selecione os dois arquivos antes de processar.")
            return
        try:
            self._status.set("Processando…")
            self.update()
            banco        = _ler_excel_banco(xp)
            erp          = _ler_txt_erp(tp)
            nao_baixadas = _comparar_duplicatas(banco, erp)

            for item in self.tree.get_children():
                self.tree.delete(item)

            nao_baixadas_set = {info['duplicata'] for info in nao_baixadas}
            total_ok  = 0
            total_nok = 0

            for norm, info in banco.items():
                if info['duplicata'] in nao_baixadas_set:
                    tag    = "nok"
                    status = "✘  Não baixada"
                    total_nok += 1
                else:
                    tag    = "ok"
                    status = "✔  Baixada"
                    total_ok += 1
                def _parse(v):
                    try:
                        return float(str(v).replace("R$","").replace(".","").replace(",",".").strip())
                    except:
                        return 0.0
                titulo  = _parse(info['valor_titulo'])
                cobrado = _parse(info['valor_cobrado'])
                juros   = cobrado - titulo
                vjuros  = "{:,.2f}".format(juros).replace(",","X").replace(".",",").replace("X",".") if juros > 0 else "-"
                self.tree.insert("", "end",
                                 values=(status, info['duplicata'], info['pagador'],
                                         info['valor_titulo'], info['valor_cobrado'], vjuros),
                                 tags=(tag,))

            if total_nok == 0:
                self._status.set(
                    f"✔  Todas as {total_ok} duplicatas estão baixadas no ERP.")
            else:
                self._status.set(
                    f"✔ {total_ok} baixada(s)   ✘ {total_nok} não baixada(s)   "
                    f"— Total: {len(banco)} no retorno bancário.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
            self._status.set(f"Erro: {e}")

    def _resetar(self):
        self._xlsx_path = None
        self._txt_path  = None
        self._var_xlsx.set("Nenhum arquivo selecionado")
        self._var_txt.set("Nenhum arquivo selecionado")
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._status.set("Selecione os arquivos e clique em Processar.")


# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 2 — SEGURO DE VIDA
# ════════════════════════════════════════════════════════════════════════════════

def _norm_nome(nome):
    """
    Normaliza nome para comparação robusta:
    - Remove acentos (João → JOAO, Ângela → ANGELA)
    - Converte para maiúsculas
    - Mantém apenas letras e espaços
    - Colapsa espaços múltiplos
    """
    s = str(nome).strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Z\s]', ' ', s)
    return ' '.join(s.split())

# Padrão de CPF — usado para delimitar o nome no PDF
_CPF_PAT = re.compile(r'\d{3}\.\d{3}\.\d{3}-\d{2}')

def _ler_pdf_seguro(path):
    """
    Extrai segurados do PDF VG Express usando CPF como delimitador.
    Retorna dict {nome_norm: {"original": str, "status": "ativo"|"inclusao"|"exclusao"}}
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber não instalado")

    resultado = {}
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                m = _CPF_PAT.search(line)
                if not m:
                    continue
                nome = line[:m.start()].strip()
                resto = line[m.end():].strip()
                # Ignora cabeçalho da tabela
                if not nome or len(nome) < 3 or 'Nome Segurado' in nome:
                    continue
                if re.search(r'xclu', resto, re.I):
                    status = "exclusao"
                elif re.search(r'nclu', resto, re.I):
                    status = "inclusao"
                else:
                    status = "ativo"
                norm = _norm_nome(nome)
                if norm:
                    resultado[norm] = {"original": nome, "status": status}
    return resultado

# Palavras que indicam células de resumo/rótulo — não são nomes de pessoas
_PALAVRAS_ROTULO = {
    'TOTAL', 'VALOR', 'QTD', 'QUANTIDADE', 'SEGURADO', 'SEGURADOS',
    'FATURA', 'PESSOA', 'PESSOAS', 'EMPRESA', 'FILIAL', 'UNIDADE',
    'DATA', 'NOME', 'CPF', 'CARGO', 'SETOR', 'DEPARTAMENTO',
}

def _parece_nome(s_norm):
    """Retorna True se a string normalizada parece um nome de pessoa."""
    words = s_norm.split()
    # Precisa ter ao menos 2 palavras, cada uma com 2+ letras
    if len(words) < 2 or any(len(w) < 2 for w in words):
        return False
    # Não pode começar com palavras típicas de rótulos
    if any(w in _PALAVRAS_ROTULO for w in words):
        return False
    return True

def _ler_excel_segurados(path, col_letra='A', start_row=1):
    """Extrai nomes de TODAS as abas do Excel. Retorna dict {nome_normalizado: nome_original}."""
    wb = openpyxl.load_workbook(path)
    col_idx = openpyxl.utils.column_index_from_string(col_letra.upper().strip())
    nomes = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if len(row) >= col_idx:
                val = row[col_idx - 1]
                if val and str(val).strip():
                    n = _norm_nome(str(val))
                    if n and _parece_nome(n):
                        nomes[n] = str(val).strip()
    return nomes


class JanelaDistribuir(tk.Toplevel):
    """Janela para distribuir novos segurados nas abas do Excel antes de salvar."""

    def __init__(self, parent, nomes, abas, xlsx_path):
        """
        nomes     : list de {"nome": str, "tipo": "inclusao"|"falta_xls"}
        abas      : list de nomes de abas disponíveis
        xlsx_path : caminho do Excel original a ser atualizado
        """
        super().__init__(parent)
        self.title("Distribuir Novos Segurados nas Abas")
        self.configure(bg=CORES["bg"])
        self.resizable(True, True)
        self.minsize(680, 460)
        self.grab_set()

        self._nomes     = sorted(nomes, key=lambda x: x["nome"])
        self._abas      = abas
        self._xlsx_path = xlsx_path
        self._combos    = {}   # {nome: StringVar}

        self._build()

    def _build(self):
        # Estilo do Combobox
        style = ttk.Style(self)
        style.configure("Dist.TCombobox",
                         fieldbackground=CORES["entry_bg"],
                         background=CORES["header"],
                         foreground=CORES["text"],
                         selectbackground=CORES["accent"],
                         selectforeground="#ffffff")
        style.map("Dist.TCombobox",
                  fieldbackground=[("readonly", CORES["entry_bg"])],
                  foreground=[("readonly", CORES["text"])])

        # Cabeçalho
        frm_top = tk.Frame(self, bg=CORES["bg"])
        frm_top.pack(fill="x", padx=16, pady=(14, 8))
        tk.Label(frm_top,
                 text=f"Selecione a aba de destino para cada novo segurado  ({len(self._nomes)} nome(s)):",
                 bg=CORES["bg"], fg=CORES["text"],
                 font=("Segoe UI", 10, "bold")).pack(side="left")
        tk.Label(frm_top,
                 text="Os nomes serão adicionados ao final da aba escolhida.",
                 bg=CORES["bg"], fg=CORES["muted"],
                 font=("Segoe UI", 8)).pack(side="left", padx=(12, 0))

        # Área scrollável
        frm_outer = tk.Frame(self, bg=CORES["bg"])
        frm_outer.pack(fill="both", expand=True, padx=16, pady=(0, 6))

        canvas = tk.Canvas(frm_outer, bg=CORES["panel"], highlightthickness=0)
        sb_y   = ttk.Scrollbar(frm_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb_y.set)
        sb_y.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=CORES["panel"])
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        # Cabeçalho da tabela
        for col, txt, w in [
            (0, "Nome",          38),
            (1, "Situação",      16),
            (2, "Aba de Destino", 20),
        ]:
            tk.Label(inner, text=txt, bg=CORES["header"], fg=CORES["accent2"],
                     font=("Segoe UI", 9, "bold"), padx=10, pady=6,
                     width=w, anchor="w" if col == 0 else "center"
                     ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

        # Linhas de dados
        for i, item in enumerate(self._nomes, start=1):
            nome = item["nome"]
            tipo = item["tipo"]
            row_bg   = CORES["entry_bg"] if i % 2 == 0 else CORES["panel"]
            tipo_txt = "✚  Inclusão"      if tipo == "inclusao" else "⚠  Sem registro"
            tipo_fg  = CORES["success"]   if tipo == "inclusao" else CORES["neutro"]

            tk.Label(inner, text=nome, bg=row_bg, fg=CORES["text"],
                     font=("Segoe UI", 9), padx=10, pady=5,
                     anchor="w").grid(row=i, column=0, sticky="nsew", padx=1, pady=1)
            tk.Label(inner, text=tipo_txt, bg=row_bg, fg=tipo_fg,
                     font=("Segoe UI", 9), padx=6, pady=5,
                     anchor="center").grid(row=i, column=1, sticky="nsew", padx=1, pady=1)

            var = tk.StringVar(value=self._abas[0] if self._abas else "")
            combo = ttk.Combobox(inner, textvariable=var, values=self._abas,
                                  style="Dist.TCombobox",
                                  state="readonly", width=20,
                                  font=("Segoe UI", 9))
            combo.grid(row=i, column=2, sticky="nsew", padx=6, pady=2)
            self._combos[nome] = var

        inner.columnconfigure(0, weight=1)

        def _on_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        def _on_canvas_resize(event):
            canvas.itemconfig(win_id, width=event.width)

        inner.bind("<Configure>", _on_resize)
        canvas.bind("<Configure>", _on_canvas_resize)

        def _scroll(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _scroll)
        self.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Rodapé
        rod = tk.Frame(self, bg=CORES["bg"])
        rod.pack(fill="x", padx=16, pady=(0, 14))

        BotaoRound(rod, "💾  Salvar Excel Atualizado", self._salvar,
                   largura=230, altura=38, raio=10
                   ).pack(side="right")
        BotaoRound(rod, "Cancelar", self.destroy,
                   largura=100, altura=38, raio=10,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"]
                   ).pack(side="right", padx=(0, 8))

    def _salvar(self):
        dest = filedialog.asksaveasfilename(
            title="Salvar Excel Atualizado",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="segurados_atualizado.xlsx",
        )
        if not dest:
            return
        try:
            wb = openpyxl.load_workbook(self._xlsx_path)
            for item in self._nomes:
                nome = item["nome"]
                aba  = self._combos[nome].get()
                if not aba:
                    continue
                if aba not in wb.sheetnames:
                    wb.create_sheet(aba)
                wb[aba].append([nome])
            wb.save(dest)
            messagebox.showinfo("Salvo",
                                f"Excel atualizado salvo com sucesso!\n\n{dest}",
                                parent=self)
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erro ao salvar", str(e), parent=self)


class FrameSeguroVida(tk.Frame):

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=CORES["bg"], **kwargs)
        self._build()

    def _build(self):
        tk.Label(self,
                 text="Compare o PDF do seguro de vida com a planilha de funcionários.",
                 bg=CORES["bg"], fg=CORES["muted"], font=("Segoe UI", 9)
                 ).pack(padx=20, pady=(14, 6), anchor="w")

        # ── Seleção de arquivos ──
        frm_f = tk.Frame(self, bg=CORES["panel"], padx=14, pady=12,
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_f.pack(fill="x", padx=20, pady=(0, 8))

        self._var_pdf  = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_xlsx = tk.StringVar(value="Nenhum arquivo selecionado")

        self._file_row(frm_f, "PDF do Seguro de Vida (.pdf):",       self._var_pdf,  self._sel_pdf,  0)
        self._file_row(frm_f, "Planilha de Funcionários (.xlsx):",   self._var_xlsx, self._sel_xlsx, 1)

        # ── Configurações da planilha ──
        frm_cfg = tk.Frame(self, bg=CORES["panel"], padx=14, pady=10,
                           highlightbackground=CORES["border"], highlightthickness=1)
        frm_cfg.pack(fill="x", padx=20, pady=(0, 8))

        tk.Label(frm_cfg, text="Configurações da planilha Excel:",
                 bg=CORES["panel"], fg=CORES["text"],
                 font=("Segoe UI", 9, "bold")).grid(row=0, column=0, columnspan=4,
                                                     sticky="w", pady=(0, 8))

        tk.Label(frm_cfg, text="Coluna dos nomes:", bg=CORES["panel"],
                 fg=CORES["text"], font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w")
        self._col_var = tk.StringVar(value="A")
        tk.Entry(frm_cfg, textvariable=self._col_var, width=5,
                 bg=CORES["entry_bg"], fg=CORES["entry_fg"], font=("Segoe UI", 9),
                 relief="flat", bd=4, insertbackground=CORES["accent"]
                 ).grid(row=1, column=1, padx=(6, 24), sticky="w")

        tk.Label(frm_cfg, text="Linha inicial:", bg=CORES["panel"],
                 fg=CORES["text"], font=("Segoe UI", 9)).grid(row=1, column=2, sticky="w")
        self._row_var = tk.StringVar(value="1")
        tk.Entry(frm_cfg, textvariable=self._row_var, width=5,
                 bg=CORES["entry_bg"], fg=CORES["entry_fg"], font=("Segoe UI", 9),
                 relief="flat", bd=4, insertbackground=CORES["accent"]
                 ).grid(row=1, column=3, padx=(6, 0), sticky="w")

        tk.Label(frm_cfg,
                 text="O PDF é lido automaticamente — os nomes são extraídos pelo CPF de cada segurado.",
                 bg=CORES["panel"], fg=CORES["muted"], font=("Segoe UI", 8)
                 ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 0))

        # ── Botões ──
        frm_btns = tk.Frame(self, bg=CORES["bg"])
        frm_btns.pack(fill="x", padx=20, pady=(0, 8))

        BotaoRound(frm_btns, "▶  Processar Comparação", self._processar,
                   largura=340, altura=40, raio=12
                   ).pack(side="left", expand=True)
        BotaoRound(frm_btns, "💾  Salvar Excel", self._abrir_distribuir,
                   largura=160, altura=40, raio=12,
                   cor=CORES["header"], cor_hover="#1a3a20",
                   cor_texto=CORES["success"],
                   fonte=("Segoe UI", 10)
                   ).pack(side="left", padx=(8, 0))
        BotaoRound(frm_btns, "🔍  Ver Extraídos", self._ver_extraidos,
                   largura=150, altura=40, raio=12,
                   cor=CORES["header"], cor_hover=CORES["accent"],
                   cor_texto=CORES["accent2"],
                   fonte=("Segoe UI", 10)
                   ).pack(side="left", padx=(8, 0))
        BotaoRound(frm_btns, "↺  Limpar", self._resetar,
                   largura=110, altura=40, raio=12,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"]
                   ).pack(side="left", padx=(8, 0))

        # ── Treeview resultado ──
        frm_r = tk.Frame(self, bg=CORES["panel"],
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_r.pack(fill="both", expand=True, padx=20, pady=(0, 6))

        style = ttk.Style()
        style.configure("Seg.Treeview",
                        background=CORES["entry_bg"], foreground=CORES["text"],
                        fieldbackground=CORES["entry_bg"],
                        font=("Segoe UI", 10), rowheight=26)
        style.configure("Seg.Treeview.Heading",
                        background=CORES["header"], foreground=CORES["accent2"],
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Seg.Treeview",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "#ffffff")])

        self.tree = ttk.Treeview(frm_r, columns=("situacao", "nome", "origem"),
                                  show="headings", style="Seg.Treeview", height=12)
        self.tree.heading("situacao", text="Situação / Ação", anchor="center")
        self.tree.heading("nome",     text="Nome",           anchor="w")
        self.tree.heading("origem",   text="Origem",         anchor="center")
        self.tree.column("situacao", width=250, anchor="center", stretch=tk.NO)
        self.tree.column("nome",     width=340, anchor="w",      stretch=tk.YES)
        self.tree.column("origem",   width=110, anchor="center", stretch=tk.NO)

        sb = ttk.Scrollbar(frm_r, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # inclusao  = verde  — adicionar ao Excel
        # exclusao  = vermelho — remover do Excel
        # falta_xls = amarelo — ativo no seguro, falta na planilha
        # falta_pdf = ciano   — na planilha, não aparece no seguro
        self.tree.tag_configure("ok",      background=CORES["success_bg"], foreground=CORES["success"])
        self.tree.tag_configure("review",  background="#fefce8",           foreground="#92400e")
        self.tree.tag_configure("so_um",   background=CORES["danger_bg"],  foreground=CORES["danger"])

        # ── Status ──
        self._status = tk.StringVar(value="Selecione os arquivos e clique em Processar.")
        tk.Label(self, textvariable=self._status,
                 font=("Segoe UI", 9), bg=CORES["bg"], fg=CORES["muted"], anchor="w"
                 ).pack(fill="x", padx=22, pady=(0, 10))

    def _file_row(self, parent, lbl, var, cmd, row):
        tk.Label(parent, text=lbl, font=("Segoe UI", 9),
                 bg=CORES["panel"], fg=CORES["text"], width=34, anchor="w"
                 ).grid(row=row, column=0, pady=6, sticky="w")
        tk.Label(parent, textvariable=var, font=("Segoe UI", 9),
                 fg=CORES["muted"], bg=CORES["panel"], width=32, anchor="w"
                 ).grid(row=row, column=1, padx=8)
        BotaoRound(parent, "Selecionar…", cmd,
                   largura=110, altura=32, raio=8,
                   cor=CORES["header"], cor_hover=CORES["accent"],
                   cor_texto=CORES["text"],
                   fonte=("Segoe UI", 9)
                   ).grid(row=row, column=2, pady=4)

    def _sel_pdf(self):
        f = filedialog.askopenfilename(title="PDF do Seguro de Vida",
                                       filetypes=[("PDF", "*.pdf")])
        if f:
            self._pdf_path = f
            self._var_pdf.set(os.path.basename(f))

    def _sel_xlsx(self):
        f = filedialog.askopenfilename(title="Planilha de Funcionários",
                                       filetypes=[("Excel", "*.xlsx *.xls")])
        if f:
            self._xlsx_path = f
            self._var_xlsx.set(os.path.basename(f))

    def _get_config(self):
        """Lê e valida configurações. Retorna (col_letra, start_row)."""
        col_letra = self._col_var.get().strip() or "A"
        start_row = int(self._row_var.get().strip() or "1")
        return col_letra, start_row

    # Cabeçalhos a ignorar na comparação
    _CABECALHOS = {_norm_nome(x) for x in [
        "NOME", "NOME DO SEGURADO", "NOME COMPLETO", "FUNCIONARIO",
        "FUNCIONÁRIO", "NOME DO FUNCIONARIO", "NOME DO FUNCIONÁRIO",
    ]}

    def _processar(self):
        pp = getattr(self, '_pdf_path',  None)
        xp = getattr(self, '_xlsx_path', None)
        if not pp or not xp:
            messagebox.showwarning("Atenção", "Selecione os dois arquivos antes de processar.")
            return
        try:
            col_letra, start_row = self._get_config()
        except ValueError:
            messagebox.showerror("Erro", "Linha inicial deve ser um número inteiro.")
            return
        try:
            self._status.set("Processando…")
            self.update()

            self._dict_pdf  = _ler_pdf_seguro(pp)
            self._dict_xlsx = _ler_excel_segurados(xp, col_letra=col_letra, start_row=start_row)

            # Remove cabeçalhos
            for cab in self._CABECALHOS:
                self._dict_pdf.pop(cab, None)
                self._dict_xlsx.pop(cab, None)

            pdf_ativos    = {k: v for k, v in self._dict_pdf.items() if v["status"] == "ativo"}
            pdf_inclusoes = {k: v for k, v in self._dict_pdf.items() if v["status"] == "inclusao"}
            pdf_exclusoes = {k: v for k, v in self._dict_pdf.items() if v["status"] == "exclusao"}
            todos_pdf     = set(self._dict_pdf.keys())

            for item in self.tree.get_children():
                self.tree.delete(item)

            n_ok = n_rev = n_so_pdf = n_so_xls = 0
            self._nomes_para_adicionar = []

            # ── Verde: ativo no PDF E presente no Excel ──────────────────────
            for k, v in sorted(pdf_ativos.items(), key=lambda x: x[1]["original"]):
                if k in self._dict_xlsx:
                    self.tree.insert("", "end", tags=("ok",),
                                     values=("✅  Consta nos dois arquivos",
                                             v["original"], "PDF e Excel"))
                    n_ok += 1

            # ── Amarelo: pendências (inclusão ou exclusão no seguro) ──────────
            for k, v in sorted(pdf_inclusoes.items(), key=lambda x: x[1]["original"]):
                self.tree.insert("", "end", tags=("review",),
                                 values=("⚠  Entrada no seguro neste período — Confirmar no Excel",
                                         v["original"], "PDF"))
                self._nomes_para_adicionar.append({"nome": v["original"], "tipo": "inclusao"})
                n_rev += 1

            for k, v in sorted(pdf_exclusoes.items(), key=lambda x: x[1]["original"]):
                self.tree.insert("", "end", tags=("review",),
                                 values=("⚠  Saída do seguro neste período — Verificar no Excel",
                                         v["original"], "PDF"))
                n_rev += 1

            # ── Vermelho: ativo no PDF mas não está no Excel ──────────────────
            for k, v in sorted(pdf_ativos.items(), key=lambda x: x[1]["original"]):
                if k not in self._dict_xlsx:
                    self.tree.insert("", "end", tags=("so_um",),
                                     values=("❌  Consta no PDF — Não consta no Excel: Adicionar no Excel",
                                             v["original"], "Somente PDF"))
                    self._nomes_para_adicionar.append({"nome": v["original"], "tipo": "falta_xls"})
                    n_so_pdf += 1

            # ── Vermelho: no Excel mas não no PDF ─────────────────────────────
            for k, nome in sorted(self._dict_xlsx.items(), key=lambda x: x[1]):
                if k not in todos_pdf:
                    self.tree.insert("", "end", tags=("so_um",),
                                     values=("❌  Consta no Excel — Não está no PDF",
                                             nome, "Somente Excel"))
                    n_so_xls += 1

            partes = [f"✅ {n_ok} em ambos"]
            if n_rev:    partes.append(f"⚠ {n_rev} para revisão")
            if n_so_pdf: partes.append(f"❌ {n_so_pdf} só no PDF")
            if n_so_xls: partes.append(f"❌ {n_so_xls} só no Excel")
            self._status.set("   |   ".join(partes))

        except ImportError:
            messagebox.showerror("Dependência faltando",
                                  "A biblioteca pdfplumber não está instalada.\n\n"
                                  "Execute o arquivo Ferramentas_Krambeck.bat que instala automaticamente,\n"
                                  "ou rode: pip install pdfplumber")
            self._status.set("Erro: pdfplumber não instalado.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
            self._status.set(f"Erro: {e}")

    def _ver_extraidos(self):
        """Abre janela mostrando os nomes brutos extraídos de cada arquivo para diagnóstico."""
        pp = getattr(self, '_pdf_path',  None)
        xp = getattr(self, '_xlsx_path', None)
        if not pp and not xp:
            messagebox.showwarning("Atenção", "Selecione ao menos um arquivo primeiro.")
            return
        try:
            col_letra, start_row = self._get_config()
        except ValueError:
            messagebox.showerror("Erro", "Configurações inválidas.")
            return

        win = tk.Toplevel(self)
        win.title("Nomes Extraídos — Diagnóstico")
        win.configure(bg=CORES["bg"])
        win.minsize(800, 500)
        win.grab_set()

        tk.Label(win,
                 text="Visualize os nomes extraídos de cada arquivo para identificar divergências de formatação.",
                 bg=CORES["bg"], fg=CORES["muted"], font=("Segoe UI", 9)
                 ).pack(padx=16, pady=(12, 6), anchor="w")

        frm = tk.Frame(win, bg=CORES["bg"])
        frm.pack(fill="both", expand=True, padx=16, pady=(0, 8))
        frm.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(1, weight=1)

        # Cabeçalhos das colunas
        for col, titulo, cor in [(0, "📄  PDF do Seguro", CORES["danger"]),
                                  (1, "📊  Excel de Funcionários", CORES["success"])]:
            tk.Label(frm, text=titulo, bg=CORES["bg"], fg=cor,
                     font=("Segoe UI", 10, "bold")).grid(row=0, column=col, sticky="w", pady=(0, 4))

        # ── Carrega ambos os dicts para poder cruzar ──────────────────────────
        d_pdf = {}
        d_xls = {}
        err_pdf = err_xls = None

        if pp:
            try:
                d_pdf = _ler_pdf_seguro(pp)
                for cab in self._CABECALHOS:
                    d_pdf.pop(cab, None)
            except ImportError:
                err_pdf = "pdfplumber não instalado."
            except Exception as e:
                err_pdf = f"Erro: {e}"

        if xp:
            try:
                d_xls = _ler_excel_segurados(xp, col_letra=col_letra, start_row=start_row)
                for cab in self._CABECALHOS:
                    d_xls.pop(cab, None)
            except Exception as e:
                err_xls = f"Erro: {e}"

        todos_pdf = set(d_pdf.keys())
        todos_xls = set(d_xls.keys())

        # ── Text widgets coloridos ─────────────────────────────────────────────
        def _make_text(parent, col):
            f = tk.Frame(parent, bg=CORES["entry_bg"],
                         highlightbackground=CORES["border"], highlightthickness=1)
            f.grid(row=1, column=col, sticky="nsew", padx=(0, 8) if col == 0 else 0)
            f.rowconfigure(0, weight=1)
            f.columnconfigure(0, weight=1)
            txt = tk.Text(f, bg=CORES["entry_bg"], fg=CORES["text"],
                          font=("Consolas", 9), relief="flat", bd=6,
                          wrap="none", state="disabled")
            sb_y = tk.Scrollbar(f, command=txt.yview, bg=CORES["panel"])
            sb_x = tk.Scrollbar(f, orient="horizontal", command=txt.xview, bg=CORES["panel"])
            txt.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
            txt.grid(row=0, column=0, sticky="nsew")
            sb_y.grid(row=0, column=1, sticky="ns")
            sb_x.grid(row=1, column=0, sticky="ew")
            txt.tag_configure("ok",     foreground=CORES["success"])
            txt.tag_configure("review", foreground="#d97706")
            txt.tag_configure("so_um",  foreground=CORES["danger"])
            txt.tag_configure("muted",  foreground=CORES["muted"])
            return txt

        txt_pdf  = _make_text(frm, 0)
        txt_xlsx = _make_text(frm, 1)

        # ── Preenche PDF ───────────────────────────────────────────────────────
        txt_pdf.config(state="normal")
        txt_pdf.delete("1.0", "end")
        if err_pdf:
            txt_pdf.insert("end", err_pdf, "muted")
        elif not d_pdf:
            txt_pdf.insert("end", "Nenhum arquivo PDF selecionado.", "muted")
        else:
            _MAP = {"ativo": "ATIVO    ", "inclusao": "ENTRADA  ", "exclusao": "SAÍDA    "}
            for norm, info in sorted(d_pdf.items(), key=lambda x: x[1]["original"]):
                status = info["status"]
                if status in ("inclusao", "exclusao"):
                    tag = "review"
                elif norm in todos_xls:
                    tag = "ok"
                else:
                    tag = "so_um"
                label = _MAP.get(status, status.upper()[:9].ljust(9))
                txt_pdf.insert("end", f"[{label}]  {info['original']}\n", tag)
        txt_pdf.config(state="disabled")

        # ── Preenche Excel ─────────────────────────────────────────────────────
        txt_xlsx.config(state="normal")
        txt_xlsx.delete("1.0", "end")
        if err_xls:
            txt_xlsx.insert("end", err_xls, "muted")
        elif not d_xls:
            txt_xlsx.insert("end", "Nenhum arquivo Excel selecionado.", "muted")
        else:
            for norm, nome in sorted(d_xls.items(), key=lambda x: x[1]):
                status_pdf = d_pdf.get(norm, {}).get("status")
                # Usa o nome do PDF como referência quando houver correspondência
                nome_ref = d_pdf[norm]["original"] if norm in d_pdf else nome
                if norm not in todos_pdf:
                    tag = "so_um"
                elif status_pdf in ("inclusao", "exclusao"):
                    tag = "review"
                else:
                    tag = "ok"
                txt_xlsx.insert("end", f"{nome_ref}\n", tag)
        txt_xlsx.config(state="disabled")

        # ── Legenda ────────────────────────────────────────────────────────────
        leg = tk.Frame(win, bg=CORES["bg"])
        leg.pack(padx=16, anchor="w")
        for cor, txt in [(CORES["success"], "Consta nos dois"),
                         ("#d97706",        "Movimentação — verificar"),
                         (CORES["danger"],  "Somente neste arquivo")]:
            tk.Label(leg, text="●", fg=cor, bg=CORES["bg"],
                     font=("Segoe UI", 11)).pack(side="left")
            tk.Label(leg, text=txt, fg=CORES["muted"], bg=CORES["bg"],
                     font=("Segoe UI", 8)).pack(side="left", padx=(0, 14))

        rod = tk.Frame(win, bg=CORES["bg"])
        rod.pack(fill="x", padx=16, pady=(6, 12))
        BotaoRound(rod, "Fechar", win.destroy,
                   largura=90, altura=32, raio=8,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"],
                   fonte=("Segoe UI", 9)
                   ).pack(side="right")

    def _abrir_distribuir(self):
        """Abre a janela para distribuir nomes novos nas abas e salvar Excel."""
        xp = getattr(self, '_xlsx_path', None)
        if not xp:
            messagebox.showwarning("Atenção", "Selecione a planilha Excel antes.", parent=self)
            return
        nomes = getattr(self, '_nomes_para_adicionar', None)
        if nomes is None:
            messagebox.showwarning("Atenção",
                                   "Processe os arquivos primeiro para identificar\n"
                                   "os segurados a adicionar.", parent=self)
            return
        if not nomes:
            messagebox.showinfo("Nada a adicionar",
                                "Não há nomes novos para adicionar ao Excel.", parent=self)
            return
        # Detecta abas disponíveis (com nomes de colaboradores)
        try:
            wb = openpyxl.load_workbook(xp)
            col_idx = openpyxl.utils.column_index_from_string(
                self._col_var.get().strip() or "A")
            abas = []
            for ws in wb.worksheets:
                tem_nome = any(
                    _parece_nome(_norm_nome(str(row[col_idx - 1])))
                    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True)
                    if len(row) >= col_idx and row[col_idx - 1]
                )
                if tem_nome:
                    abas.append(ws.title)
            if not abas:
                abas = [ws.title for ws in wb.worksheets]
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível ler as abas do Excel:\n{e}", parent=self)
            return
        JanelaDistribuir(self, nomes, abas, xp)

    def _resetar(self):
        self._pdf_path  = None
        self._xlsx_path = None
        self._nomes_para_adicionar = []
        self._var_pdf.set("Nenhum arquivo selecionado")
        self._var_xlsx.set("Nenhum arquivo selecionado")
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._status.set("Selecione os arquivos e clique em Processar.")


# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 3 — CONTAS A PAGAR
# ════════════════════════════════════════════════════════════════════════════════

WEBHOOK_URL = ""
CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")

EMPRESAS = ["L01", "L02", "L03", "L04", "L05"]
BANCOS   = ["VIACREDI", "BRADESCO", "SANTANDER", "ITAU"]

TEMPLATE_PADRAO = """\
📋 **Resumo Diário — {DATA}**

💼 **Pagamentos do Dia**
L01 | Pag: {L01_PAG} | Saldo: {L01_SALDO} | Dif: {L01_DIF}
L02 | Pag: {L02_PAG} | Saldo: {L02_SALDO} | Dif: {L02_DIF}
L03 | Pag: {L03_PAG} | Saldo: {L03_SALDO} | Dif: {L03_DIF}
L04 | Pag: {L04_PAG} | Saldo: {L04_SALDO} | Dif: {L04_DIF}
L05 | Pag: {L05_PAG} | Saldo: {L05_SALDO} | Dif: {L05_DIF}

🏦 **Saldos por Banco**
L01 | VIACREDI: {L01_VIACREDI} | BRADESCO: {L01_BRADESCO} | SANTANDER: {L01_SANTANDER} | ITAU: {L01_ITAU}
L02 | VIACREDI: {L02_VIACREDI} | BRADESCO: {L02_BRADESCO} | SANTANDER: {L02_SANTANDER} | ITAU: {L02_ITAU}
L03 | VIACREDI: {L03_VIACREDI} | BRADESCO: {L03_BRADESCO} | SANTANDER: {L03_SANTANDER} | ITAU: {L03_ITAU}
L04 | VIACREDI: {L04_VIACREDI} | BRADESCO: {L04_BRADESCO} | SANTANDER: {L04_SANTANDER} | ITAU: {L04_ITAU}
L05 | VIACREDI: {L05_VIACREDI} | BRADESCO: {L05_BRADESCO} | SANTANDER: {L05_SANTANDER} | ITAU: {L05_ITAU}
"""

VARIAVEIS_DISPONIVEIS = [
    ("{DATA}", "Data de referência"),
] + [
    ("{%s_PAG}" % e,   "%s — Pagamentos" % e)        for e in EMPRESAS
] + [
    ("{%s_SALDO}" % e, "%s — Saldo total" % e)        for e in EMPRESAS
] + [
    ("{%s_DIF}" % e,   "%s — Diferença (saldo - pag)" % e) for e in EMPRESAS
] + [
    ("{%s_%s}" % (e, b), "%s — %s" % (e, b))
    for e in EMPRESAS for b in BANCOS
]

# ── Helpers ──────────────────────────────────────────────────────────────────

def format_brl(v):
    return "R$ {:,.2f}".format(v).replace(",","X").replace(".",",").replace("X",".")

def parse_brl(t):
    try:
        return float(t.replace("R$","").replace(".","").replace(",",".").strip())
    except:
        return 0.0

# ── Persistência ─────────────────────────────────────────────────────────────

def salvar_config(webhook, template, pagamentos, saldos):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "webhook":    webhook,
            "template":   template,
            "pagamentos": pagamentos,
            "saldos":     saldos,
        }, f, ensure_ascii=False, indent=2)

def carregar_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"webhook": "", "template": TEMPLATE_PADRAO, "pagamentos": {}, "saldos": {}}

# ── Discord ───────────────────────────────────────────────────────────────────

def _post_discord(url, payload):
    data = json.dumps(payload).encode("utf-8")
    req  = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json", "User-Agent": "ContasPagarApp/1.0"},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            status = resp.status
            if status not in (200, 204):
                raise ConnectionError("Discord retornou status inesperado: {}".format(status))
            return status
    except ConnectionError:
        raise
    except urllib.error.HTTPError as e:
        raise ConnectionError(_msg_erro(e.code))
    except urllib.error.URLError as e:
        raise ConnectionError(
            "Sem conexao com o Discord.\nVerifique sua internet.\n\nDetalhe: {}".format(e.reason))
    except Exception as e:
        raise ConnectionError("Erro inesperado ao enviar: {}".format(e))

def _msg_erro(code):
    return {
        400: "400 - Requisicao invalida. O formato pode estar incorreto.",
        401: "401 - Nao autorizado. A URL do webhook e invalida.",
        403: ("403 - Acesso negado.\n\n"
              "Verifique:\n"
              "  1. A URL esta completa e correta?\n"
              "  2. O webhook ainda existe no canal?\n"
              "  3. Tente deletar e criar um novo webhook."),
        404: "404 - Webhook nao encontrado. Pode ter sido deletado.",
        429: "429 - Muitas requisicoes. Aguarde e tente novamente.",
    }.get(code, "Erro HTTP {}.".format(code))

def construir_variaveis(pagamentos, saldos, data_ref):
    ctx = {"DATA": data_ref}
    for emp in EMPRESAS:
        pag   = pagamentos.get(emp, 0.0)
        saldo = sum(saldos.get(emp, {}).get(b, 0.0) for b in BANCOS)
        dif   = saldo - pag
        ctx["{}_PAG".format(emp)]   = format_brl(pag)
        ctx["{}_SALDO".format(emp)] = format_brl(saldo)
        ctx["{}_DIF".format(emp)]   = format_brl(dif)
        for b in BANCOS:
            ctx["{}_{}".format(emp, b)] = format_brl(saldos.get(emp, {}).get(b, 0.0))
    return ctx

def renderizar_template(template, ctx):
    msg = template
    for k, v in ctx.items():
        msg = msg.replace("{" + k + "}", v)
    return msg

def enviar_discord(url, template, pagamentos, saldos, data_ref):
    if not url.strip():
        raise ValueError("URL do webhook nao configurada.")
    if not (url.startswith("https://discord.com/api/webhooks/") or
            url.startswith("https://discordapp.com/api/webhooks/")):
        raise ValueError("URL invalida.\nFormato: https://discord.com/api/webhooks/ID/TOKEN")
    ctx      = construir_variaveis(pagamentos, saldos, data_ref)
    mensagem = renderizar_template(template, ctx)
    partes   = [mensagem[i:i+2000] for i in range(0, len(mensagem), 2000)]
    for parte in partes:
        _post_discord(url, {"username": "Contas a Pagar", "content": parte})

def testar_webhook(url):
    if not url.strip():
        raise ValueError("URL do webhook nao configurada.")
    _post_discord(url, {"username": "Contas a Pagar", "content": "Teste de conexao OK!"})

# ── Janela de Configurações (Toplevel) ───────────────────────────────────────

class JanelaConfig(tk.Toplevel):
    def __init__(self, parent, webhook_var, template_var):
        super().__init__(parent)
        self.title("Configuracoes")
        self.configure(bg=CORES["bg"])
        self.resizable(True, True)
        self.minsize(620, 560)
        self.grab_set()

        style = ttk.Style(self)
        style.configure("Cfg.TNotebook", background=CORES["bg"], borderwidth=0)
        style.configure("Cfg.TNotebook.Tab", background=CORES["panel"],
                        foreground=CORES["muted"], font=("Segoe UI", 9), padding=[12, 5])
        style.map("Cfg.TNotebook.Tab",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "white")])
        nb = ttk.Notebook(self, style="Cfg.TNotebook")
        nb.pack(fill="both", expand=True, padx=12, pady=12)

        aba_wh = tk.Frame(nb, bg=CORES["bg"])
        nb.add(aba_wh, text="  Webhook  ")
        self._build_aba_webhook(aba_wh, webhook_var)

        aba_msg = tk.Frame(nb, bg=CORES["bg"])
        nb.add(aba_msg, text="  Mensagem  ")
        self._build_aba_mensagem(aba_msg, template_var)

        rod = tk.Frame(self, bg=CORES["bg"])
        rod.pack(fill="x", padx=12, pady=(0, 12))
        BotaoRound(rod, "Fechar", self.destroy,
                   largura=90, altura=32, raio=8,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"],
                   fonte=("Segoe UI", 9)
                   ).pack(side="right")

    def _build_aba_webhook(self, parent, webhook_var):
        tk.Label(parent, text="URL do Webhook Discord",
                 bg=CORES["bg"], fg=CORES["text"],
                 font=("Segoe UI", 10, "bold")).pack(padx=14, pady=(14, 2), anchor="w")
        tk.Label(parent,
                 text="Discord > Canal > Editar canal > Integracoes > Webhooks > Copiar URL",
                 bg=CORES["bg"], fg=CORES["muted"],
                 font=("Segoe UI", 8)).pack(padx=14, anchor="w")

        self._wh_var = tk.StringVar(value=webhook_var.get())
        tk.Entry(parent, textvariable=self._wh_var, width=68,
                 bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                 insertbackground=CORES["accent"],
                 relief="flat", bd=6, font=("Segoe UI", 9)
                 ).pack(padx=14, pady=8, fill="x")

        self._wh_status = tk.Label(parent, text="", bg=CORES["bg"],
                                   fg=CORES["muted"], font=("Segoe UI", 8),
                                   wraplength=500, justify="left")
        self._wh_status.pack(padx=14, anchor="w")

        frm = tk.Frame(parent, bg=CORES["bg"])
        frm.pack(padx=14, pady=10, fill="x")
        BotaoRound(frm, "Salvar URL",
                   lambda: self._salvar_webhook(webhook_var),
                   largura=120, altura=34, raio=8,
                   fonte=("Segoe UI", 9, "bold")
                   ).pack(side="right")
        BotaoRound(frm, "Testar conexão", self._testar,
                   largura=140, altura=34, raio=8,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["text"],
                   fonte=("Segoe UI", 9)
                   ).pack(side="right", padx=(0, 8))

    def _testar(self):
        try:
            testar_webhook(self._wh_var.get().strip())
            self._wh_status.config(text="Conexao OK! Mensagem de teste enviada.", fg=CORES["success"])
        except Exception as e:
            self._wh_status.config(text="Erro: {}".format(e), fg=CORES["danger"])

    def _salvar_webhook(self, wv):
        wv.set(self._wh_var.get().strip())
        self._wh_status.config(text="URL salva.", fg=CORES["success"])

    def _build_aba_mensagem(self, parent, template_var):
        parent.columnconfigure(0, weight=3)
        parent.columnconfigure(1, weight=1, minsize=200)
        parent.rowconfigure(0, weight=1)

        left = tk.Frame(parent, bg=CORES["bg"])
        left.grid(row=0, column=0, sticky="nsew", padx=(8, 4), pady=8)
        left.rowconfigure(1, weight=1)
        left.columnconfigure(0, weight=1)

        tk.Label(left, text="Template da mensagem",
                 bg=CORES["bg"], fg=CORES["text"],
                 font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 2))

        txt_frame = tk.Frame(left, bg=CORES["entry_bg"])
        txt_frame.grid(row=1, column=0, sticky="nsew")
        txt_frame.rowconfigure(0, weight=1)
        txt_frame.columnconfigure(0, weight=1)

        self._txt = tk.Text(txt_frame, bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                            insertbackground=CORES["accent"],
                            font=("Consolas", 9), relief="flat", bd=6,
                            wrap="word", undo=True)
        sb = tk.Scrollbar(txt_frame, command=self._txt.yview,
                          bg=CORES["panel"], troughcolor=CORES["bg"])
        self._txt.configure(yscrollcommand=sb.set)
        self._txt.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        self._txt.insert("1.0", template_var.get())

        btn_frm = tk.Frame(left, bg=CORES["bg"])
        btn_frm.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        BotaoRound(btn_frm, "Salvar template",
                   lambda: self._salvar_template(template_var),
                   largura=140, altura=32, raio=8,
                   fonte=("Segoe UI", 9, "bold")
                   ).pack(side="right")
        BotaoRound(btn_frm, "Restaurar padrão", self._restaurar,
                   largura=150, altura=32, raio=8,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"],
                   fonte=("Segoe UI", 9)
                   ).pack(side="right", padx=(0, 8))

        self._tmpl_status = tk.Label(left, text="", bg=CORES["bg"],
                                     fg=CORES["muted"], font=("Segoe UI", 8))
        self._tmpl_status.grid(row=3, column=0, sticky="w", pady=(4, 0))

        right = tk.Frame(parent, bg=CORES["bg"])
        right.grid(row=0, column=1, sticky="nsew", padx=(4, 8), pady=8)
        right.rowconfigure(1, weight=1)
        right.columnconfigure(0, weight=1)

        tk.Label(right, text="Variaveis disponiveis",
                 bg=CORES["bg"], fg=CORES["text"],
                 font=("Segoe UI", 9, "bold")).grid(row=0, column=0, columnspan=2,
                                                    sticky="w", pady=(0, 2))

        vsb = tk.Scrollbar(right, bg=CORES["panel"], troughcolor=CORES["bg"])
        vsb.grid(row=1, column=1, sticky="ns")
        listbox = tk.Listbox(right, bg=CORES["entry_bg"], fg=CORES["text"],
                             selectbackground=CORES["accent"],
                             font=("Consolas", 8), relief="flat", bd=0,
                             yscrollcommand=vsb.set, activestyle="none",
                             highlightthickness=0)
        vsb.config(command=listbox.yview)
        listbox.grid(row=1, column=0, sticky="nsew")

        for var, desc in VARIAVEIS_DISPONIVEIS:
            listbox.insert("end", "  {}".format(var))

        self._tip = tk.Label(right, text="Clique para inserir no cursor",
                             bg=CORES["header"], fg=CORES["muted"],
                             font=("Segoe UI", 8), padx=6, pady=4,
                             wraplength=180, justify="left")
        self._tip.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(4, 0))

        def on_select(event):
            sel = listbox.curselection()
            if not sel:
                return
            var, desc = VARIAVEIS_DISPONIVEIS[sel[0]]
            self._tip.config(text=desc, fg=CORES["text"])
            try:
                self._txt.insert(tk.INSERT, var)
                self._txt.focus_set()
            except:
                pass

        listbox.bind("<<ListboxSelect>>", on_select)

    def _salvar_template(self, tv):
        tv.set(self._txt.get("1.0", "end-1c"))
        self._tmpl_status.config(text="Template salvo.", fg=CORES["success"])

    def _restaurar(self):
        self._txt.delete("1.0", "end")
        self._txt.insert("1.0", TEMPLATE_PADRAO)
        self._tmpl_status.config(text="Template restaurado para o padrao.", fg=CORES["muted"])


# ── Tabela Editável ───────────────────────────────────────────────────────────

class TabelaEditavel(tk.Frame):
    def __init__(self, parent, headers, rows, readonly=None, **kwargs):
        super().__init__(parent, bg=CORES["panel"], **kwargs)
        self.headers  = headers
        self.row_keys = rows
        self.readonly = readonly or []
        self._cells   = {}
        for c, h in enumerate(headers):
            tk.Label(self, text=h, bg=CORES["header"], fg=CORES["accent2"],
                     font=("Segoe UI", 9, "bold"), padx=10, pady=6,
                     anchor="center").grid(row=0, column=c, sticky="nsew", padx=1, pady=1)
        for r, rk in enumerate(rows, start=1):
            for c, h in enumerate(headers):
                var = tk.StringVar()
                self._cells[(rk, h)] = var
                if c == 0:
                    tk.Label(self, textvariable=var, bg=CORES["panel"],
                             fg=CORES["text"], font=("Segoe UI", 9, "bold"),
                             padx=10, pady=4, anchor="center").grid(
                                 row=r, column=c, sticky="nsew", padx=1, pady=1)
                    var.set(rk)
                elif c in self.readonly:
                    tk.Label(self, textvariable=var, bg=CORES["bg"],
                             fg=CORES["neutro"], font=("Segoe UI", 9),
                             padx=8, pady=4, anchor="center").grid(
                                 row=r, column=c, sticky="nsew", padx=1, pady=1)
                    var.set("R$ 0,00")
                else:
                    ent = tk.Entry(self, textvariable=var, justify="center",
                                   bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                                   insertbackground=CORES["accent"],
                                   font=("Segoe UI", 9), relief="flat", bd=4)
                    ent.grid(row=r, column=c, sticky="nsew", padx=1, pady=1)
                    ent.bind("<FocusIn>",  lambda e, v=var: self._fi(v))
                    ent.bind("<FocusOut>", lambda e, v=var: self._fo(v))
                    var.set("R$ 0,00")
        for c in range(len(headers)):
            self.columnconfigure(c, weight=1, minsize=110)

    def _fi(self, var):
        val = var.get().replace("R$", "").replace(".", "").replace(",", ".").strip()
        try:
            var.set("{:.2f}".format(float(val)).replace(".", ","))
        except:
            var.set("0,00")

    def _fo(self, var):
        try:
            var.set(format_brl(float(var.get().replace(",", ".").strip())))
        except:
            var.set("R$ 0,00")

    def get_value(self, rk, h):
        v = self._cells.get((rk, h))
        return parse_brl(v.get()) if v else 0.0

    def set_value(self, rk, h, valor):
        v = self._cells.get((rk, h))
        if v:
            v.set(format_brl(valor))


# ── Frame Contas a Pagar ──────────────────────────────────────────────────────

class FrameContasPagar(tk.Frame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=CORES["bg"], **kwargs)
        cfg = carregar_config()
        self._webhook  = tk.StringVar(value=cfg.get("webhook", WEBHOOK_URL))
        self._template = tk.StringVar(value=cfg.get("template", TEMPLATE_PADRAO))
        self._build_ui()
        self._carregar_dados(cfg)
        self._atualizar_resumo()

    def _build_ui(self):
        # Barra de data + config
        topo = tk.Frame(self, bg=CORES["header"], height=50)
        topo.pack(fill="x")
        topo.pack_propagate(False)

        df = tk.Frame(topo, bg=CORES["header"])
        df.pack(side="left", padx=16, pady=10)
        tk.Label(df, text="Data de referencia:", bg=CORES["header"],
                 fg=CORES["muted"], font=("Segoe UI", 9)).pack(side="left")
        self._data_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        tk.Entry(df, textvariable=self._data_var, width=12,
                 bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                 font=("Segoe UI", 9), relief="flat", bd=4,
                 insertbackground=CORES["accent"]).pack(side="left", padx=8)

        BotaoRound(topo, "⚙  Webhook / Mensagem", self._abrir_config,
                   largura=190, altura=32, raio=8,
                   cor=CORES["panel"], cor_hover=CORES["accent"],
                   cor_texto=CORES["muted"],
                   fonte=("Segoe UI", 9)
                   ).pack(side="right", padx=12, pady=9)

        # Notebook
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("D.TNotebook", background=CORES["bg"], borderwidth=0)
        style.configure("D.TNotebook.Tab", background=CORES["panel"],
                        foreground=CORES["muted"], font=("Segoe UI", 10), padding=[16, 6])
        style.map("D.TNotebook.Tab",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "white")])
        nb = ttk.Notebook(self, style="D.TNotebook")
        nb.pack(fill="both", expand=True, padx=16, pady=10)

        a1 = tk.Frame(nb, bg=CORES["bg"])
        nb.add(a1, text="  Viacredi")
        self._build_viacredi(a1)

        a2 = tk.Frame(nb, bg=CORES["bg"])
        nb.add(a2, text="  Saldos por Banco")
        self._build_bancos(a2)

        # Rodapé
        rod = tk.Frame(self, bg=CORES["header"], height=60)
        rod.pack(fill="x", side="bottom")
        rod.pack_propagate(False)

        btn_frame = tk.Frame(rod, bg=CORES["header"])
        btn_frame.pack(side="right", padx=16, pady=11)

        BotaoRound(btn_frame, "Enviar para Discord", self._enviar_discord,
                   largura=200, altura=38, raio=10
                   ).pack(side="right")
        BotaoRound(btn_frame, "Salvar", self._salvar,
                   largura=100, altura=38, raio=10,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["text"]
                   ).pack(side="right", padx=(0, 8))
        BotaoRound(btn_frame, "Recalcular", self._atualizar_resumo,
                   largura=120, altura=38, raio=10,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["text"]
                   ).pack(side="right", padx=(0, 8))

    def _build_viacredi(self, parent):
        self._vc_saldo_var = {}
        self._vc_pag_var   = {}

        tk.Label(parent,
                 text="Informe o saldo atual e os pagamentos Viacredi por loja.",
                 bg=CORES["bg"], fg=CORES["muted"],
                 font=("Segoe UI", 9)).pack(padx=12, pady=(8, 4), anchor="w")

        frm = tk.Frame(parent, bg=CORES["panel"])
        frm.pack(fill="both", expand=True, padx=12, pady=8)

        COLS = ["LOJA", "SALDO VIACREDI", "PAGAMENTOS"]
        for c, h in enumerate(COLS):
            tk.Label(frm, text=h, bg=CORES["header"], fg=CORES["accent2"],
                     font=("Segoe UI", 9, "bold"), padx=10, pady=6,
                     anchor="center").grid(row=0, column=c, sticky="nsew", padx=1, pady=1)

        for r, emp in enumerate(EMPRESAS, start=1):
            # LOJA
            tk.Label(frm, text=emp, bg=CORES["panel"],
                     fg=CORES["text"], font=("Segoe UI", 9, "bold"),
                     padx=10, pady=4, anchor="center").grid(
                         row=r, column=0, sticky="nsew", padx=1, pady=1)

            # SALDO VIACREDI
            sv = tk.StringVar(value="R$ 0,00")
            self._vc_saldo_var[emp] = sv
            ent_s = tk.Entry(frm, textvariable=sv, justify="center",
                             bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                             insertbackground=CORES["accent"],
                             font=("Segoe UI", 9), relief="flat", bd=4)
            ent_s.grid(row=r, column=1, sticky="nsew", padx=1, pady=1)
            ent_s.bind("<FocusIn>",  lambda e, v=sv: self._vc_fi(v))
            ent_s.bind("<FocusOut>", lambda e, v=sv: self._vc_fo(v))

            # PAGAMENTOS
            pv = tk.StringVar(value="R$ 0,00")
            self._vc_pag_var[emp] = pv
            ent_p = tk.Entry(frm, textvariable=pv, justify="center",
                             bg=CORES["entry_bg"], fg=CORES["entry_fg"],
                             insertbackground=CORES["accent"],
                             font=("Segoe UI", 9), relief="flat", bd=4)
            ent_p.grid(row=r, column=2, sticky="nsew", padx=1, pady=1)
            ent_p.bind("<FocusIn>",  lambda e, v=pv: self._vc_fi(v))
            ent_p.bind("<FocusOut>", lambda e, v=pv: self._vc_fo(v))

        for c in range(len(COLS)):
            frm.columnconfigure(c, weight=1, minsize=130)

    def _vc_fi(self, var):
        val = var.get().replace("R$", "").replace(".", "").replace(",", ".").strip()
        try:
            var.set("{:.2f}".format(float(val)).replace(".", ","))
        except Exception:
            var.set("0,00")

    def _vc_fo(self, var):
        try:
            var.set(format_brl(float(var.get().replace(",", ".").strip())))
        except Exception:
            var.set("R$ 0,00")


    def _build_bancos(self, parent):
        tk.Label(parent, text="Informe o saldo atual em cada conta bancaria.",
                 bg=CORES["bg"], fg=CORES["muted"],
                 font=("Segoe UI", 9)).pack(padx=12, pady=(8, 4), anchor="w")
        bancos_sem_vc = [b for b in BANCOS if b != "VIACREDI"]
        self._tbl_bancos = TabelaEditavel(
            parent, headers=["LOJA", "DATA"] + bancos_sem_vc, rows=EMPRESAS, readonly=[1])
        self._tbl_bancos.pack(fill="both", expand=True, padx=12, pady=8)
        for emp in EMPRESAS:
            v = self._tbl_bancos._cells.get((emp, "DATA"))
            if v:
                v.set(datetime.now().strftime("%d/%m/%Y"))

    def _atualizar_resumo(self):
        ds = self._data_var.get()
        for emp in EMPRESAS:
            v2 = self._tbl_bancos._cells.get((emp, "DATA"))
            if v2:
                v2.set(ds)

    def _coletar_pagamentos(self):
        return {emp: parse_brl(self._vc_pag_var[emp].get()) for emp in EMPRESAS}

    def _coletar_saldos(self):
        bancos_sem_vc = [b for b in BANCOS if b != "VIACREDI"]
        return {emp: dict(
            {b: self._tbl_bancos.get_value(emp, b) for b in bancos_sem_vc},
            VIACREDI=parse_brl(self._vc_saldo_var[emp].get())
        ) for emp in EMPRESAS}

    def _salvar(self):
        salvar_config(self._webhook.get(), self._template.get(),
                      self._coletar_pagamentos(), self._coletar_saldos())
        messagebox.showinfo("Salvo", "Dados salvos!", parent=self)

    def _enviar_discord(self):
        self._atualizar_resumo()
        url     = self._webhook.get().strip()
        enviado = False
        erro    = None
        if not url:
            messagebox.showwarning("Webhook", "Configure a URL do webhook primeiro.", parent=self)
            return
        try:
            enviar_discord(url, self._template.get(),
                           self._coletar_pagamentos(), self._coletar_saldos(),
                           self._data_var.get())
            enviado = True
        except Exception as e:
            erro = str(e)
        if enviado:
            messagebox.showinfo("Enviado!", "Mensagem enviada ao Discord com sucesso!", parent=self)
        else:
            messagebox.showerror("Erro ao enviar", erro or "Erro desconhecido.", parent=self)

    def _abrir_config(self):
        JanelaConfig(self, self._webhook, self._template)

    def _carregar_dados(self, cfg):
        pags          = cfg.get("pagamentos", {})
        saldos        = cfg.get("saldos", {})
        bancos_sem_vc = [b for b in BANCOS if b != "VIACREDI"]
        for emp in EMPRESAS:
            # Viacredi
            self._vc_saldo_var[emp].set(format_brl(saldos.get(emp, {}).get("VIACREDI", 0.0)))
            self._vc_pag_var[emp].set(format_brl(pags.get(emp, 0.0)))
            # Outros bancos
            for b in bancos_sem_vc:
                self._tbl_bancos.set_value(emp, b, saldos.get(emp, {}).get(b, 0.0))


# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 4 — VERIFICADOR DE CRÉDITOS EM ABERTO x BAIXADOS
# ════════════════════════════════════════════════════════════════════════════════

def _br_float(s):
    return round(float(str(s).replace(".", "").replace(",", ".")), 2)


def _parse_data(s):
    """Converte strings DD/MM/YYYY, YYYY-MM-DD ou objetos datetime para date.
    Retorna None se não for possível fazer o parse."""
    if not s:
        return None
    if hasattr(s, "date"):          # objeto datetime do openpyxl
        return s.date() if hasattr(s, "hour") else s
    s = str(s).strip()[:10]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parsear_txt_baixados(arq_txt):
    """Lê um relatório TXT de duplicatas baixadas e retorna estruturas de lookup."""
    with open(arq_txt, encoding="latin-1") as f:
        txt = f.read()

    periodo = re.search(r'Pagamento de\s+(\d{2}/\d{2}/\d{4})\s+at[eé]\s+(\d{2}/\d{2}/\d{4})', txt)
    if not periodo:
        raise ValueError("Não foi possível identificar o período no arquivo TXT.")

    dt_ini, dt_fim = periodo.group(1), periodo.group(2)
    ano = dt_fim[-4:]

    pessoa_re = r'Pessoa:\s+(\d+)'
    trans_re  = (r'^\s+(\d+)\s+(\S+)\s+(\S*)\s+'
                 r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+'
                 r'(-?\d+)\s+([\d.,]+)\s+([\d.,]+)')

    baixados     = {}
    detalhes_txt = {}
    current      = None

    for line in txt.split('\n'):
        pm = re.search(pessoa_re, line)
        if pm:
            current = pm.group(1).strip()
            baixados.setdefault(current, set())
        elif current:
            tm = re.match(trans_re, line)
            if tm:
                try:
                    val = _br_float(tm.group(8))
                    mov = _br_float(tm.group(9))
                    baixados[current].add(val)
                    key = (current, val)
                    detalhes_txt.setdefault(key, []).append({
                        "duplicata":   tm.group(2),
                        "emissao":     tm.group(4),
                        "vencimento":  tm.group(5),
                        "pagamento":   tm.group(6),
                        "valor":       val,
                        "valor_mov":   mov,
                    })
                except ValueError:
                    pass

    todos_baixados = {v for vs in baixados.values() for v in vs}
    return baixados, detalhes_txt, todos_baixados, ano, dt_ini, dt_fim


def _analisar_creditos(arq_txt, arq_xlsx, arq_desconto=None):
    baixados_cr, detalhes_cr, todos_cr, ano, dt_ini, dt_fim = _parsear_txt_baixados(arq_txt)

    # Desconto direto (opcional)
    if arq_desconto:
        baixados_dc, detalhes_dc, todos_dc, _, _, _ = _parsear_txt_baixados(arq_desconto)
    else:
        baixados_dc, detalhes_dc, todos_dc = {}, {}, set()

    todos_baixados = todos_cr | todos_dc

    wb = openpyxl.load_workbook(arq_xlsx)
    if ano not in wb.sheetnames:
        raise ValueError(f"Aba '{ano}' não encontrada. Abas disponíveis: {', '.join(wb.sheetnames)}")

    ws = wb[ano]
    aberto = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[4] is None or not isinstance(row[4], (int, float)):
            continue
        aberto.append({
            "emissao": str(row[0]) if row[0] else "",
            "cod":     str(row[2]).strip() if row[2] else "",
            "forn":    str(row[3]) if row[3] else "",
            "valor":   round(float(row[4]), 2),
            "saldo":   round(float(row[5]), 2) if isinstance(row[5], (int, float)) else 0,
        })

    def _validos(detalhes_src, cod, valor, dt_emissao):
        todos_det = detalhes_src.get((cod, valor), [])
        if not todos_det:
            return []
        if dt_emissao:
            return [d for d in todos_det
                    if _parse_data(d["pagamento"]) is None
                    or dt_emissao <= _parse_data(d["pagamento"])]
        return todos_det

    confirmados, suspeitos = [], []
    for cr in aberto:
        cod, valor   = cr["cod"], cr["valor"]
        dt_emissao   = _parse_data(cr["emissao"])
        adicionado   = False

        # 1. Verifica baixa como CRÉDITO
        if cod in baixados_cr and valor in baixados_cr[cod]:
            v = _validos(detalhes_cr, cod, valor, dt_emissao)
            if v:
                entry = dict(cr, match="CONFIRMADO", tipo_baixa="Crédito", detalhes=v)
                confirmados.append(entry)
                adicionado = True

        # 2. Verifica baixa como DESCONTO DIRETO
        if cod in baixados_dc and valor in baixados_dc[cod]:
            v = _validos(detalhes_dc, cod, valor, dt_emissao)
            if v:
                entry = dict(cr, match="CONFIRMADO", tipo_baixa="Desconto Direto", detalhes=v)
                confirmados.append(entry)
                adicionado = True

        if adicionado:
            continue

        # 3. Suspeito: mesmo valor, fornecedor diferente
        if valor in todos_baixados:
            if dt_emissao:
                algum_valido = any(
                    _parse_data(d["pagamento"]) is None or dt_emissao <= _parse_data(d["pagamento"])
                    for src in (detalhes_cr, detalhes_dc)
                    for dets in src.values()
                    for d in dets
                    if d["valor"] == valor
                )
                if not algum_valido:
                    continue
            cr["match"]    = "SUSPEITO"
            cr["detalhes"] = []
            suspeitos.append(cr)

    return {
        "ano": ano, "dt_ini": dt_ini, "dt_fim": dt_fim,
        "total_aberto": len(aberto),
        "confirmados":  confirmados,
        "suspeitos":    suspeitos,
    }


class FrameVerificadorCreditos(tk.Frame):
    """
    Módulo: Créditos em Aberto × Baixados
    Cruza a planilha de créditos em aberto (garantia) com o relatório de
    baixados exportado do ERP e identifica divergências.
    """

    # Resultado atual (para exportação)
    _resultado = None

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=CORES["bg"], **kwargs)
        self._txt_path  = None
        self._xlsx_path = None
        self._build()
        self._auto_preencher()

    # ── Preenchimento automático ──────────────────────────────────────────────

    def _auto_preencher(self):
        pasta = os.path.dirname(os.path.abspath(__file__))
        txt  = os.path.join(pasta, "CREDITOS BAIXADOS.txt")
        xlsx = os.path.join(pasta, "CREDITOS EM ABERTO.xlsx")
        if os.path.exists(txt):
            self._txt_path = txt
            self._var_txt.set(os.path.basename(txt))
        if os.path.exists(xlsx):
            self._xlsx_path = xlsx
            self._var_xlsx.set(os.path.basename(xlsx))

    # ── Construção da interface ───────────────────────────────────────────────

    def _build(self):

        # ── 1. Cabeçalho descritivo ──
        cab = tk.Frame(self, bg=CORES["header"],
                       highlightbackground=CORES["border"], highlightthickness=1)
        cab.pack(fill="x", padx=20, pady=(14, 0))

        tk.Label(cab, text="  🔍  Verificador de Créditos em Aberto",
                 font=("Segoe UI", 11, "bold"), bg=CORES["header"],
                 fg=CORES["text"]).pack(side="left", padx=4, pady=10)
        tk.Label(cab,
                 text="Identifica créditos que constam em aberto na planilha da garantia "
                      "mas já foram baixados no sistema ERP.",
                 font=("Segoe UI", 8), bg=CORES["header"],
                 fg=CORES["muted"]).pack(side="left", padx=(0, 10), pady=10)

        # ── 2. Seleção de arquivos ──
        frm_f = tk.Frame(self, bg=CORES["panel"], padx=14, pady=10,
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_f.pack(fill="x", padx=20, pady=(0, 0))

        self._var_txt      = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_xlsx     = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_desconto = tk.StringVar(value="Nenhum arquivo selecionado")

        self._file_row(frm_f,
                       "📄  Baixados — ERP (.txt)",
                       "Relatório exportado do sistema com as duplicatas pagas no período.",
                       self._var_txt, self._sel_txt, 0)
        self._file_row(frm_f,
                       "📊  Em Aberto — Garantia (.xlsx)",
                       "Planilha do setor de garantia com os créditos de fornecedores em aberto.",
                       self._var_xlsx, self._sel_xlsx, 1)
        self._file_row(frm_f,
                       "🔖  Desconto Direto (.txt)  —  opcional",
                       "Relatório de baixas por desconto direto. Se informado, identifica créditos fechados pela condição errada.",
                       self._var_desconto, self._sel_desconto, 2)

        # ── 3. Botões de ação ──
        frm_btns = tk.Frame(self, bg=CORES["bg"])
        frm_btns.pack(fill="x", padx=20, pady=8)

        self._btn = BotaoRound(frm_btns, "▶  Verificar Créditos", self._executar,
                               largura=480, altura=38, raio=10)
        self._btn.pack(side="left", expand=True)

        self._btn_exp = BotaoRound(frm_btns, "📥  Exportar Excel", self._exportar,
                                   largura=160, altura=38, raio=10,
                                   cor=CORES["panel"], cor_hover="#1a3a20",
                                   cor_texto=CORES["success"])
        self._btn_exp.pack(side="left", padx=(8, 0))

        BotaoRound(frm_btns, "↺  Limpar", self._resetar,
                   largura=110, altura=38, raio=10,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"]
                   ).pack(side="left", padx=(8, 0))

        # ── 4. Cards de resumo ──
        self._frm_cards = tk.Frame(self, bg=CORES["bg"])
        self._frm_cards.pack(fill="x", padx=20, pady=(0, 8))
        self._cards = {}
        card_defs = [
            ("total",  "📋 Em Aberto",     CORES["muted"],   "Total de créditos\nna planilha"),
            ("conf",   "✅ Confirmados",    CORES["success"], "Mesmo fornecedor\ne mesmo valor no ERP"),
            ("susp",   "⚠️  Suspeitos",     CORES["neutro"],  "Mesmo valor, mas\nfornecedor diferente"),
            ("ok",     "✔  Sem Pendência",  CORES["accent2"], "Sem correspondência\nnos baixados"),
        ]
        for key, titulo, cor, descricao in card_defs:
            card = tk.Frame(self._frm_cards, bg=CORES["panel"],
                            highlightbackground=cor, highlightthickness=1)
            card.pack(side="left", expand=True, fill="x", padx=(0, 8), pady=2)
            tk.Label(card, text=titulo, font=("Segoe UI", 8, "bold"),
                     bg=CORES["panel"], fg=cor).pack(padx=10, pady=(8, 2), anchor="w")
            val_lbl = tk.Label(card, text="—", font=("Segoe UI", 18, "bold"),
                               bg=CORES["panel"], fg=cor)
            val_lbl.pack(padx=10, pady=(0, 2), anchor="w")
            tk.Label(card, text=descricao, font=("Segoe UI", 7),
                     bg=CORES["panel"], fg=CORES["muted"], justify="left"
                     ).pack(padx=10, pady=(0, 8), anchor="w")
            self._cards[key] = val_lbl

        # ── 5. Notebook de resultados ──
        style = ttk.Style()
        style.configure("Cred.TNotebook",     background=CORES["bg"], borderwidth=0)
        style.configure("Cred.TNotebook.Tab", background=CORES["panel"],
                        foreground=CORES["muted"], font=("Segoe UI", 9), padding=[12, 5])
        style.map("Cred.TNotebook.Tab",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "white")])

        self._nb = ttk.Notebook(self, style="Cred.TNotebook")
        self._nb.pack(fill="both", expand=True, padx=20, pady=(0, 4))

        # Tab Confirmados
        self._tab_conf = tk.Frame(self._nb, bg=CORES["bg"])
        self._nb.add(self._tab_conf, text="✅  Confirmados")
        tk.Label(self._tab_conf,
                 text="  Créditos confirmados como já baixados no ERP — mesmo fornecedor e mesmo valor. "
                      "Verifique se o registro em aberto deve ser encerrado.",
                 font=("Segoe UI", 8), bg=CORES["bg"], fg=CORES["muted"]
                 ).pack(anchor="w", padx=4, pady=(6, 2))
        self._tree_conf = self._criar_tree(self._tab_conf,
            cols=("Data de Emissão", "Fornecedor", "Valor em Aberto (R$)",
                  "Valor Baixado (R$)", "Nº Duplicata", "Data da Baixa", "Tipo de Baixa"),
            widths=(130, 270, 150, 150, 140, 130, 130),
            tag="conf")

        # Tab Suspeitos
        self._tab_susp = tk.Frame(self._nb, bg=CORES["bg"])
        self._nb.add(self._tab_susp, text="⚠️   Suspeitos")
        tk.Label(self._tab_susp,
                 text="  Créditos com valor igual ao de alguma duplicata baixada, mas de fornecedor diferente. "
                      "Pode ser coincidência — verifique manualmente.",
                 font=("Segoe UI", 8), bg=CORES["bg"], fg=CORES["muted"]
                 ).pack(anchor="w", padx=4, pady=(6, 2))
        self._tree_susp = self._criar_tree(self._tab_susp,
            cols=("Data de Emissão", "Fornecedor", "Valor em Aberto (R$)", "Saldo (R$)"),
            widths=(130, 420, 160, 130),
            tag="susp")

        # Tab Resumo
        self._tab_res = tk.Frame(self._nb, bg=CORES["bg"])
        self._nb.add(self._tab_res, text="📋  Resumo")
        self._txt_res = self._criar_text(self._tab_res)

        # ── 6. Barra de status ──
        self._status = tk.StringVar(value="Selecione os arquivos e clique em Verificar Créditos.")
        tk.Label(self, textvariable=self._status,
                 font=("Segoe UI", 8), bg=CORES["bg"], fg=CORES["muted"], anchor="w"
                 ).pack(fill="x", padx=22, pady=(0, 6))

    # ── Helpers de widget ─────────────────────────────────────────────────────

    def _file_row(self, parent, titulo, dica, var, cmd, row):
        tk.Label(parent, text=titulo, font=("Segoe UI", 9, "bold"),
                 bg=CORES["panel"], fg=CORES["text"], anchor="w", width=32
                 ).grid(row=row*2,   column=0, padx=(0, 8), pady=(6, 0), sticky="w")
        tk.Label(parent, text=dica, font=("Segoe UI", 7),
                 bg=CORES["panel"], fg=CORES["muted"], anchor="w"
                 ).grid(row=row*2+1, column=0, padx=(0, 8), pady=(0, 6), sticky="w")
        tk.Label(parent, textvariable=var, font=("Segoe UI", 9),
                 fg=CORES["accent2"], bg=CORES["panel"], anchor="w", width=34
                 ).grid(row=row*2, column=1, rowspan=2, padx=8, sticky="w")
        BotaoRound(parent, "Selecionar…", cmd,
                   largura=110, altura=30, raio=8,
                   cor=CORES["header"], cor_hover=CORES["accent"],
                   cor_texto=CORES["text"], fonte=("Segoe UI", 9)
                   ).grid(row=row*2, column=2, rowspan=2, pady=4)

    def _criar_tree(self, parent, cols, widths, tag):
        style = ttk.Style()
        sname = f"CV_{tag}.Treeview"
        style.configure(sname,
                        background=CORES["entry_bg"], foreground=CORES["text"],
                        fieldbackground=CORES["entry_bg"],
                        font=("Segoe UI", 9), rowheight=24)
        style.configure(sname + ".Heading",
                        background=CORES["header"], foreground=CORES["accent2"],
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map(sname, background=[("selected", CORES["accent"])],
                  foreground=[("selected", "#fff")])

        frm = tk.Frame(parent, bg=CORES["panel"],
                       highlightbackground=CORES["border"], highlightthickness=1)
        frm.pack(fill="both", expand=True, padx=0, pady=(0, 0))

        sb_y = ttk.Scrollbar(frm, orient="vertical")
        sb_x = ttk.Scrollbar(frm, orient="horizontal")
        tree = ttk.Treeview(frm, columns=cols, show="headings", style=sname,
                            yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.config(command=tree.yview)
        sb_x.config(command=tree.xview)

        for i, col in enumerate(cols):
            w = widths[i] if widths else 110
            is_valor = any(x in col for x in ("Valor", "Saldo"))
            anc = "e" if is_valor else "w"
            # heading anchor deve bater com o anchor dos dados
            tree.heading(col, text=col, anchor=anc)
            # Fornecedor (maior coluna de texto) recebe stretch; demais são fixas
            stretch = tk.YES if ("Fornecedor" in col or "Nome" in col or "Pagador" in col) else tk.NO
            tree.column(col, width=w, anchor=anc, minwidth=60, stretch=stretch)

        cor_fg = CORES["success"] if tag == "conf" else CORES["neutro"]
        cor_bg = CORES["success_bg"] if tag == "conf" else "#fefce8"  # amarelo-claro para suspeitos
        cor_fg_susp = "#92400e"  # âmbar escuro para leitura sobre fundo amarelo
        if tag != "conf":
            cor_fg = cor_fg_susp
        tree.tag_configure("main",     background=cor_bg,         foreground=cor_fg)
        tree.tag_configure("alt",      background=CORES["panel"], foreground=CORES["text"])
        tree.tag_configure("desconto", background="#fff7ed",      foreground="#c2410c")

        sb_y.pack(side="right",  fill="y")
        sb_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)
        return tree

    def _criar_text(self, parent):
        frm = tk.Frame(parent, bg=CORES["panel"],
                       highlightbackground=CORES["border"], highlightthickness=1)
        frm.pack(fill="both", expand=True)
        sb = tk.Scrollbar(frm, bg=CORES["panel"])
        sb.pack(side="right", fill="y")
        txt = tk.Text(frm, font=("Consolas", 9), bg=CORES["entry_bg"], fg=CORES["text"],
                      insertbackground=CORES["accent"], relief="flat", bd=10,
                      yscrollcommand=sb.set, state="disabled", wrap="none")
        txt.pack(fill="both", expand=True)
        sb.config(command=txt.yview)
        txt.tag_configure("h1",     foreground=CORES["accent"],  font=("Segoe UI", 12, "bold"))
        txt.tag_configure("h2",     foreground=CORES["accent2"], font=("Segoe UI", 9,  "bold"))
        txt.tag_configure("ok",     foreground=CORES["success"])
        txt.tag_configure("warn",   foreground=CORES["neutro"])
        txt.tag_configure("info",   foreground=CORES["accent2"])
        txt.tag_configure("muted",  foreground=CORES["muted"])
        txt.tag_configure("normal", foreground=CORES["text"])
        return txt

    def _sel_txt(self):
        f = filedialog.askopenfilename(title="Créditos Baixados — ERP",
                                       filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if f:
            self._txt_path = f
            self._var_txt.set(os.path.basename(f))

    def _sel_desconto(self):
        f = filedialog.askopenfilename(title="Desconto Direto — ERP",
                                       filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if f:
            self._desconto_path = f
            self._var_desconto.set(os.path.basename(f))

    def _sel_xlsx(self):
        f = filedialog.askopenfilename(title="Créditos em Aberto — Garantia",
                                       filetypes=[("Excel", "*.xlsx")])
        if f:
            self._xlsx_path = f
            self._var_xlsx.set(os.path.basename(f))

    # ── Execução ──────────────────────────────────────────────────────────────

    def _executar(self):
        if not self._txt_path or not self._xlsx_path:
            messagebox.showwarning("Atenção", "Selecione os dois arquivos antes de verificar.")
            return
        self._btn.config_text("⏳  Analisando…")
        self.update()
        try:
            r = _analisar_creditos(self._txt_path, self._xlsx_path,
                                   getattr(self, "_desconto_path", None))
            self._resultado = r
            conf = len(r["confirmados"])
            susp = len(r["suspeitos"])
            sem  = r["total_aberto"] - conf - susp

            # Atualiza cards
            self._cards["total"].config(text=str(r["total_aberto"]))
            self._cards["conf"].config(text=str(conf))
            self._cards["susp"].config(text=str(susp))
            self._cards["ok"].config(text=str(sem))

            self._preencher_confirmados(r["confirmados"])
            self._preencher_suspeitos(r["suspeitos"])
            self._preencher_resumo(r)

            if conf:
                self._nb.select(0)
            elif susp:
                self._nb.select(1)
            else:
                self._nb.select(2)

            total_r = sum(c["valor"] for c in r["confirmados"])
            self._status.set(
                f"Ano {r['ano']}  |  ✅ {conf} confirmado(s)  R$ {total_r:,.2f}  |  "
                f"⚠️ {susp} suspeito(s)  |  ✔ {sem} sem pendência")
        except Exception as e:
            messagebox.showerror("Erro na análise", str(e))
            self._status.set(f"Erro: {e}")
        finally:
            self._btn.config_text("▶  Verificar Créditos")

    def _preencher_confirmados(self, confirmados):
        t = self._tree_conf
        t.delete(*t.get_children())
        for cr in confirmados:
            tipo = cr.get("tipo_baixa", "Crédito")
            tag  = "desconto" if tipo == "Desconto Direto" else "main"
            for det in (cr["detalhes"] or [{}]):
                t.insert("", "end", tags=(tag,), values=(
                    cr["emissao"],
                    cr["forn"],
                    f"R$ {cr['valor']:,.2f}",
                    f"R$ {det.get('valor_mov', cr['valor']):,.2f}",
                    det.get("duplicata", ""),
                    det.get("pagamento", ""),
                    tipo,
                ))
        self._nb.tab(0, text=f"✅  Confirmados ({len(confirmados)})")

    def _preencher_suspeitos(self, suspeitos):
        t = self._tree_susp
        t.delete(*t.get_children())
        for i, cr in enumerate(suspeitos):
            t.insert("", "end", tags=("main" if i % 2 == 0 else "alt",), values=(
                cr["emissao"],
                cr["forn"],
                f"R$ {cr['valor']:,.2f}",
                f"R$ {cr['saldo']:,.2f}",
            ))
        self._nb.tab(1, text=f"⚠️   Suspeitos ({len(suspeitos)})")

    def _preencher_resumo(self, r):
        conf = r["confirmados"]
        susp = r["suspeitos"]
        sem  = r["total_aberto"] - len(conf) - len(susp)
        total_r = sum(c["valor"] for c in conf)

        t = self._txt_res
        t.config(state="normal")
        t.delete("1.0", "end")

        t.insert("end", f"\n  RELATÓRIO DE CRÉDITOS EM ABERTO — {r['ano']}\n\n", "h1")
        t.insert("end", f"  Período analisado : {r['dt_ini']} a {r['dt_fim']}\n", "muted")
        t.insert("end", f"  Arquivo ERP       : {os.path.basename(self._txt_path)}\n", "muted")
        t.insert("end", f"  Planilha Garantia : {os.path.basename(self._xlsx_path)}\n\n", "muted")

        t.insert("end", "  RESUMO\n", "h2")
        t.insert("end",  "  " + "─" * 60 + "\n", "muted")
        t.insert("end", f"  Total de créditos em aberto          : {r['total_aberto']}\n", "normal")
        t.insert("end", f"  ✅  Confirmados (baixados no ERP)     : {len(conf)}   →   R$ {total_r:,.2f}\n", "ok")
        t.insert("end", f"  ⚠️   Suspeitos (só valor coincide)     : {len(susp)}\n", "warn")
        t.insert("end", f"  ✔   Sem correspondência               : {sem}\n\n", "muted")

        if not conf and not susp:
            t.insert("end", "  ✔  Nenhum crédito em aberto encontrado nos baixados do ERP.\n", "ok")
        else:
            if conf:
                t.insert("end", "\n  CONFIRMADOS — DETALHES PARA PESQUISA NO SISTEMA\n", "h2")
                t.insert("end",  "  " + "─" * 60 + "\n", "muted")
                t.insert("end",
                         f"  {'Data Crédito':<14} {'Cód':>6}  {'Fornecedor':<38}"
                         f"  {'Valor':>11}  {'Nº Duplicata':<22}"
                         f"  {'Emissão Dup':<12}  {'Data de Baixa'}\n", "h2")
                t.insert("end", "  " + "─" * 120 + "\n", "muted")
                for cr in conf:
                    for d in (cr["detalhes"] or [{}]):
                        t.insert("end",
                                 f"  {cr['emissao']:<14} {cr['cod']:>6}  {cr['forn']:<38}"
                                 f"  R$ {cr['valor']:>8,.2f}  {d.get('duplicata',''):<22}"
                                 f"  {d.get('emissao',''):<12}  {d.get('pagamento','')}\n", "normal")
                t.insert("end", "\n")

            if susp:
                t.insert("end", "\n  SUSPEITOS — VERIFICAR MANUALMENTE\n", "h2")
                t.insert("end",  "  " + "─" * 60 + "\n", "muted")
                t.insert("end",
                         f"  {'Data Crédito':<14} {'Cód':>6}  {'Fornecedor':<38}  {'Valor':>11}\n", "h2")
                t.insert("end", "  " + "─" * 80 + "\n", "muted")
                for cr in susp:
                    t.insert("end",
                             f"  {cr['emissao']:<14} {cr['cod']:>6}  {cr['forn']:<38}"
                             f"  R$ {cr['valor']:>8,.2f}\n", "warn")

        t.config(state="disabled")

    # ── Exportação ────────────────────────────────────────────────────────────

    def _exportar(self):
        if not self._resultado:
            messagebox.showwarning("Atenção", "Execute a verificação antes de exportar.")
            return
        r = self._resultado
        dest = filedialog.asksaveasfilename(
                        title="Salvar resultado",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="creditos_pendentes_" + r["ano"] + ".xlsx",
        )
        if not dest:
            return
        try:
            wb  = openpyxl.Workbook()
            hdr = {"font":      openpyxl.styles.Font(bold=True, color="FFFFFF"),
                   "fill":      openpyxl.styles.PatternFill("solid", fgColor="1A1A38"),
                   "alignment": openpyxl.styles.Alignment(horizontal="center")}

            # Aba Confirmados
            ws = wb.active
            ws.title = "Confirmados " + r["ano"]
            cabecalho = ["Data do Credito", "Cod. Fornecedor", "Fornecedor",
                         "Valor do Credito (R$)", "Saldo em Aberto (R$)",
                         "Nr Duplicata ERP", "Emissao Dup.", "Vencimento", "Data de Baixa"]
            ws.append(cabecalho)
            for cell in ws[1]:
                cell.font      = hdr["font"]
                cell.fill      = hdr["fill"]
                cell.alignment = hdr["alignment"]
            for cr in r["confirmados"]:
                for d in (cr["detalhes"] or [{}]):
                    ws.append([
                        cr["emissao"], cr["cod"], cr["forn"],
                        cr["valor"], cr["saldo"],
                        d.get("duplicata", ""), d.get("emissao", ""),
                        d.get("vencimento", ""), d.get("pagamento", ""),
                    ])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(
                    len(str(c.value or "")) for c in col) + 4

            # Aba Suspeitos
            if r["suspeitos"]:
                ws2 = wb.create_sheet("Suspeitos " + r["ano"])
                cab2 = ["Data do Credito", "Cod. Fornecedor", "Fornecedor",
                        "Valor do Credito (R$)", "Saldo em Aberto (R$)"]
                ws2.append(cab2)
                for cell in ws2[1]:
                    cell.font      = hdr["font"]
                    cell.fill      = hdr["fill"]
                    cell.alignment = hdr["alignment"]
                for cr in r["suspeitos"]:
                    ws2.append([cr["emissao"], cr["cod"], cr["forn"],
                                cr["valor"], cr["saldo"]])
                for col in ws2.columns:
                    ws2.column_dimensions[col[0].column_letter].width = max(
                        len(str(c.value or "")) for c in col) + 4

            wb.save(dest)
            messagebox.showinfo("Exportado", "Arquivo salvo com sucesso!\n\n" + dest)
        except Exception as e:
            messagebox.showerror("Erro ao exportar", str(e))

    # ── Reset ─────────────────────────────────────────────────────────────────

    def _resetar(self):
        self._txt_path      = None
        self._xlsx_path     = None
        self._desconto_path = None
        self._resultado     = None
        self._var_txt.set("Nenhum arquivo selecionado")
        self._var_xlsx.set("Nenhum arquivo selecionado")
        self._var_desconto.set("Nenhum arquivo selecionado")
        for key in self._cards:
            self._cards[key].config(text="—")
        self._tree_conf.delete(*self._tree_conf.get_children())
        self._tree_susp.delete(*self._tree_susp.get_children())
        self._txt_res.config(state="normal")
        self._txt_res.delete("1.0", "end")
        self._txt_res.config(state="disabled")
        self._nb.tab(0, text="✅  Confirmados")
        self._nb.tab(1, text="⚠️   Suspeitos")
        self._status.set("Selecione os arquivos e clique em Verificar Creditos.")


# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 5 — COMPARADOR DDA × CONTAS A PAGAR
# ════════════════════════════════════════════════════════════════════════════════

def _dda_parse_br(s):
    return float(s.replace('R$', '').strip().replace('.', '').replace(',', '.') or '0')

def _dda_fmt_br(v):
    return "R$ " + f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def _dda_fmt_cnpj(c):
    c = re.sub(r'\D', '', c)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else c

_DDA_STOPS = {
    'DE', 'DO', 'DA', 'DOS', 'DAS', 'E', 'COM', 'LTDA', 'SA', 'ME',
    'EIRELI', 'EIRELE', 'IMP', 'EXP', 'DISTRIB', 'INDUSTRIA',
    'COMERCIO', 'COMERCIAL', 'BRASIL',
}
_DDA_FIDC_KW = {'FIDC', 'BANCO', 'FUNDO', 'CRED', 'CREDIT'}

def _dda_sig_tokens(name):
    return {t for t in re.findall(r'\b[A-Z]{4,}\b', name.upper()) if t not in _DDA_STOPS}

def _dda_carregar_cp(path):
    lines = []
    for enc in ('latin-1', 'cp1252', 'utf-8'):
        try:
            lines = open(path, encoding=enc).readlines()
            break
        except Exception:
            continue
    entries = []
    for line in lines:
        if not re.match(r'^\s{1,4}\d\s+\d+\s+\S', line):
            continue
        vm = re.search(r'([\d.]+,\d{2})\s+0,00\s+R\$', line)
        nm = re.match(r'^\s+\d+\s+\d+\s{2,}(.+?)\s{2,}(\S+)', line)
        if not vm or not nm:
            continue
        entries.append({
            'nome':      nm.group(1).strip(),
            'duplicata': nm.group(2).strip(),
            'valor':     _dda_parse_br(vm.group(1)),
            'valor_str': f"R$ {vm.group(1)}",
            'matched':   False,
        })
    return entries

def _dda_carregar_dda(path):
    entries = []
    with open(path, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f, delimiter=';'):
            entries.append({
                'beneficiario': row['Beneficiario'].strip(),
                'cnpj':         _dda_fmt_cnpj(row.get('Inscricao Benef.', '')),
                'emissao':      row.get('Emissao', ''),
                'vencimento':   row.get('Vencimento', ''),
                'nosso_numero': row.get('Nosso Numero', ''),
                'valor':        _dda_parse_br(row.get('Valor Nominal', '0')),
                'valor_str':    row.get('Valor Nominal', ''),
                'matched':      False,
                'cp_match':     None,
            })
    return entries

def _dda_comparar(cp_path, dda_path):
    cp  = _dda_carregar_cp(cp_path)
    dda = _dda_carregar_dda(dda_path)
    for d in dda:
        for c in cp:
            if c['matched']:
                continue
            if abs(c['valor'] - d['valor']) < 0.01:
                d['matched'] = True
                d['cp_match'] = c
                c['matched'] = True
                break
    sem_match, cedido_alta, cedido_verif, near = [], [], [], []
    used_cp_idx = set()
    for d in dda:
        if not d['matched']:
            sem_match.append(d)
        else:
            c = d['cp_match']
            has_fidc = bool(set(re.findall(r'\b[A-Z]+\b', d['beneficiario'].upper())) & _DDA_FIDC_KW)
            td, tc = _dda_sig_tokens(d['beneficiario']), _dda_sig_tokens(c['nome'])
            if has_fidc:
                cedido_alta.append((d, c))
            elif td and tc and not (td & tc):
                cedido_verif.append((d, c))
    for d in dda:
        if d['matched']:
            continue
        for j, c in enumerate(cp):
            if c['matched'] or j in used_cp_idx:
                continue
            diff = abs(d['valor'] - c['valor'])
            if 0.01 < diff <= 5.0:
                near.append((d, c, diff))
                used_cp_idx.add(j)
                break
    return dict(dda=dda, cp=cp, sem_match=sem_match,
                cedido_alta=cedido_alta, cedido_verif=cedido_verif, near=near)


class FrameComparadorDDA(tk.Frame):
    """Módulo: Comparador DDA × Contas a Pagar — visão unificada."""

    # Status labels — exibidos na tabela e no Excel
    _ST = {
        "ok":      "CONCILIADO",
        "sem":     "DUPLICATA SOMENTE NO DDA",
        "benef":   "CONSTA NOS DOIS, PORÉM COM BENEFICIÁRIOS DIFERENTES",
        "near":    "CONSTA NOS DOIS, PORÉM COM VALORES DIFERENTES",
        "only_cp": "DUPLICATA SOMENTE NO CONTAS A PAGAR",
    }

    # Cores de fundo Excel (hex sem #)
    _XL_FILL = {
        "ok":      "DCFCE7",
        "sem":     "FEE2E2",
        "benef":   "FEF3C7",
        "near":    "EDE9FE",
        "only_cp": "DBEAFE",
    }

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=CORES["bg"], **kwargs)
        self._resultado = None
        self._cp_path   = None
        self._dda_path  = None
        self._filtro    = "todos"
        self._all_rows  = []   # lista de (vals_tuple, tag)
        self._build()
        self._auto_detectar()

    # ── Auto-detecção ────────────────────────────────────────────────────────

    def _auto_detectar(self):
        base = os.path.dirname(os.path.abspath(__file__))
        cp   = os.path.join(base, 'CONTAS A PAGAR.txt')
        ddas = sorted(glob.glob(os.path.join(base, 'RelatorioDDA*.csv')))
        if os.path.exists(cp):
            self._cp_path = cp
            self._var_cp.set(os.path.basename(cp))
        if ddas:
            self._dda_path = ddas[-1]
            self._var_dda.set(os.path.basename(ddas[-1]))
        if self._cp_path and self._dda_path:
            self._status.set("Arquivos detectados. Clique em Comparar.")

    # ── Construção da interface ──────────────────────────────────────────────

    def _build(self):
        tk.Label(self,
                 text="Cada boleto DDA cruzado com sua duplicata no Contas a Pagar — visão completa.",
                 bg=CORES["bg"], fg=CORES["muted"], font=("Segoe UI", 9)
                 ).pack(padx=20, pady=(14, 6), anchor="w")

        # ── Arquivos
        frm_f = tk.Frame(self, bg=CORES["panel"], padx=14, pady=10,
                         highlightbackground=CORES["border"], highlightthickness=1)
        frm_f.pack(fill="x", padx=20, pady=(0, 8))
        self._var_cp  = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_dda = tk.StringVar(value="Nenhum arquivo selecionado")
        self._file_row(frm_f, "Contas a Pagar (.txt):", self._var_cp,  self._sel_cp,  0)
        self._file_row(frm_f, "Relatório DDA (.csv):",  self._var_dda, self._sel_dda, 1)

        # ── Botões de ação
        frm_btns = tk.Frame(self, bg=CORES["bg"])
        frm_btns.pack(fill="x", padx=20, pady=(0, 8))
        self._btn_comp = BotaoRound(frm_btns, "▶  Comparar", self._executar,
                                    largura=380, altura=40, raio=12)
        self._btn_comp.pack(side="left")
        BotaoRound(frm_btns, "📥  Exportar Excel", self._exportar,
                   largura=175, altura=40, raio=12,
                   cor=CORES["panel"], cor_hover="#d1fae5",
                   cor_texto=CORES["success"]
                   ).pack(side="left", padx=(8, 0))
        BotaoRound(frm_btns, "↺  Limpar", self._resetar,
                   largura=110, altura=40, raio=12,
                   cor=CORES["panel"], cor_hover=CORES["border"],
                   cor_texto=CORES["muted"]
                   ).pack(side="left", padx=(8, 0))

        # ── KPI cards
        frm_kpi = tk.Frame(self, bg=CORES["bg"])
        frm_kpi.pack(fill="x", padx=20, pady=(0, 6))
        self._kpi_vars = {}
        kpi_defs = [
            ("total",   "Total DDA",               CORES["panel"],  CORES["text"]),
            ("concil",  "Conciliados",              "#dcfce7",       "#16a34a"),
            ("sem",     "Somente no DDA",           "#fee2e2",       "#dc2626"),
            ("benef",   "Beneficiários diferentes", "#fef3c7",       "#d97706"),
            ("near",    "Valores diferentes",       "#ede9fe",       "#7c3aed"),
            ("only_cp", "Somente no CP",            "#dbeafe",       "#1d4ed8"),
        ]
        for key, lbl, cbg, cfg in kpi_defs:
            frm = tk.Frame(frm_kpi, bg=cbg, padx=10, pady=6,
                           highlightbackground=CORES["border"], highlightthickness=1)
            frm.pack(side="left", expand=True, fill="x", padx=(0, 4))
            v = tk.StringVar(value="—")
            self._kpi_vars[key] = v
            tk.Label(frm, textvariable=v, font=("Segoe UI", 16, "bold"),
                     bg=cbg, fg=cfg).pack()
            tk.Label(frm, text=lbl, font=("Segoe UI", 7),
                     bg=cbg, fg=cfg).pack()

        # ── Filtros
        frm_fil = tk.Frame(self, bg=CORES["bg"])
        frm_fil.pack(fill="x", padx=20, pady=(0, 4))
        tk.Label(frm_fil, text="Filtro:", font=("Segoe UI", 9),
                 bg=CORES["bg"], fg=CORES["muted"]).pack(side="left", padx=(0, 6))
        self._filtro_btns = {}
        _flt = [
            ("todos",   "Todos",            72),
            ("ok",      "✅ Conciliado",    115),
            ("sem",     "Somente no DDA",   120),
            ("benef",   "Benef. diferentes",130),
            ("near",    "Valores dif.",     100),
            ("only_cp", "Somente no CP",    115),
        ]
        for cod, lbl, w in _flt:
            ativo = (cod == "todos")
            btn = BotaoRound(frm_fil, lbl, lambda c=cod: self._aplicar_filtro(c),
                             largura=w, altura=28, raio=8,
                             cor=CORES["accent"] if ativo else CORES["panel"],
                             cor_hover=CORES["accent"],
                             cor_texto="#ffffff" if ativo else CORES["text"],
                             fonte=("Segoe UI", 8))
            btn.pack(side="left", padx=(0, 4))
            self._filtro_btns[cod] = btn

        # ── Treeview principal
        cols = ('Status', 'DDA — Beneficiário', 'CNPJ', 'Nosso Número', 'Vencimento',
                'DDA Valor', 'CP — Fornecedor', 'CP — Duplicata', 'CP Valor', 'Observação')
        wds  = (250, 200, 155, 130, 90, 100, 190, 100, 100, 200)

        style = ttk.Style()
        style.configure("DDA2.Treeview",
                        background=CORES["entry_bg"], foreground=CORES["text"],
                        fieldbackground=CORES["entry_bg"],
                        font=("Segoe UI", 9), rowheight=26)
        style.configure("DDA2.Treeview.Heading",
                        background=CORES["header"], foreground=CORES["accent2"],
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("DDA2.Treeview",
                  background=[("selected", CORES["accent"])],
                  foreground=[("selected", "#ffffff")])

        frm_tv = tk.Frame(self, bg=CORES["panel"],
                          highlightbackground=CORES["border"], highlightthickness=1)
        frm_tv.pack(fill="both", expand=True, padx=20, pady=(0, 4))

        sb_y = ttk.Scrollbar(frm_tv, orient="vertical")
        sb_x = ttk.Scrollbar(frm_tv, orient="horizontal")
        self._tree = ttk.Treeview(frm_tv, columns=cols, show="headings",
                                   style="DDA2.Treeview",
                                   yscrollcommand=sb_y.set,
                                   xscrollcommand=sb_x.set,
                                   height=16)
        sb_y.config(command=self._tree.yview)
        sb_x.config(command=self._tree.xview)

        for i, col in enumerate(cols):
            self._tree.heading(col, text=col, anchor="w")
            self._tree.column(col, width=wds[i], anchor="w", minwidth=50)

        self._tree.tag_configure("ok",      background="#f0fdf4")
        self._tree.tag_configure("sem",     background="#fef2f2")
        self._tree.tag_configure("benef",   background="#fffbeb")
        self._tree.tag_configure("near",    background="#f5f3ff")
        self._tree.tag_configure("only_cp", background="#eff6ff")

        sb_y.pack(side="right",  fill="y")
        sb_x.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        # ── Status
        self._status = tk.StringVar(value="Selecione os arquivos e clique em Comparar.")
        tk.Label(self, textvariable=self._status,
                 font=("Segoe UI", 9), bg=CORES["bg"], fg=CORES["muted"], anchor="w"
                 ).pack(fill="x", padx=22, pady=(0, 10))

    # ── Helpers UI ───────────────────────────────────────────────────────────

    def _file_row(self, parent, lbl, var, cmd, row):
        tk.Label(parent, text=lbl, font=("Segoe UI", 9),
                 bg=CORES["panel"], fg=CORES["text"], width=26, anchor="w"
                 ).grid(row=row, column=0, pady=4, sticky="w")
        tk.Label(parent, textvariable=var, font=("Segoe UI", 9),
                 fg=CORES["muted"], bg=CORES["panel"], width=42, anchor="w"
                 ).grid(row=row, column=1, padx=8)
        BotaoRound(parent, "Selecionar…", cmd,
                   largura=110, altura=30, raio=8,
                   cor=CORES["header"], cor_hover=CORES["accent"],
                   cor_texto=CORES["text"], fonte=("Segoe UI", 9)
                   ).grid(row=row, column=2, pady=3)

    # ── Seleção de arquivos ──────────────────────────────────────────────────

    def _sel_cp(self):
        f = filedialog.askopenfilename(
            title="Contas a Pagar",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if f:
            self._cp_path = f
            self._var_cp.set(os.path.basename(f))
            if not self._dda_path:
                ddas = sorted(glob.glob(os.path.join(os.path.dirname(f), 'RelatorioDDA*.csv')))
                if ddas:
                    self._dda_path = ddas[-1]
                    self._var_dda.set(os.path.basename(ddas[-1]))

    def _sel_dda(self):
        f = filedialog.askopenfilename(
            title="Relatório DDA",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if f:
            self._dda_path = f
            self._var_dda.set(os.path.basename(f))

    # ── Execução ─────────────────────────────────────────────────────────────

    def _executar(self):
        if not self._cp_path or not os.path.exists(self._cp_path):
            messagebox.showerror("Arquivo não encontrado",
                                 "Selecione o arquivo CONTAS A PAGAR.txt")
            return
        if not self._dda_path or not os.path.exists(self._dda_path):
            messagebox.showerror("Arquivo não encontrado",
                                 "Selecione o Relatório DDA (.csv)")
            return
        self._btn_comp.config_text("⏳  Processando…")
        self.update()
        try:
            self._resultado = _dda_comparar(self._cp_path, self._dda_path)
        except Exception as ex:
            messagebox.showerror("Erro ao processar", str(ex))
            self._btn_comp.config_text("▶  Comparar")
            return
        self._btn_comp.config_text("▶  Comparar")
        self._popular()

    def _popular(self):
        r   = self._resultado
        dda = r['dda']

        fidc_ids  = {id(d) for d, _ in r['cedido_alta']}
        verif_ids = {id(d) for d, _ in r['cedido_verif']}
        near_map  = {id(x[0]): x for x in r['near']}
        near_cp_ids = {id(x[1]) for x in r['near']}

        rows = []

        for d in dda:
            c = d.get('cp_match')
            if d['matched'] and c:
                if id(d) in fidc_ids or id(d) in verif_ids:
                    tag = "benef"
                    obs = f"Cessão/FIDC" if id(d) in fidc_ids else f"CP: {c['nome']}"
                else:
                    tag = "ok"
                    obs = ""
                cp_forn = c['nome']
                cp_dup  = c['duplicata']
                cp_val  = c['valor_str']
            else:
                nr = near_map.get(id(d))
                if nr:
                    tag     = "near"
                    obs     = f"Diferença: {_dda_fmt_br(nr[2])}"
                    cp_forn = nr[1]['nome']
                    cp_dup  = nr[1]['duplicata']
                    cp_val  = nr[1]['valor_str']
                else:
                    tag     = "sem"
                    obs     = ""
                    cp_forn = cp_dup = cp_val = ""

            rows.append((
                (self._ST[tag], d['beneficiario'], d['cnpj'], d['nosso_numero'],
                 d['vencimento'], d['valor_str'], cp_forn, cp_dup, cp_val, obs),
                tag
            ))

        for c in r['cp']:
            if not c['matched'] and id(c) not in near_cp_ids:
                rows.append((
                    (self._ST["only_cp"], "", "", "", "", "",
                     c['nome'], c['duplicata'], c['valor_str'], ""),
                    "only_cp"
                ))

        self._all_rows = rows

        counts = {t: sum(1 for _, tg in rows if tg == t)
                  for t in ("ok", "sem", "benef", "near", "only_cp")}

        self._kpi_vars["total"].set(str(len(dda)))
        self._kpi_vars["concil"].set(str(counts["ok"]))
        self._kpi_vars["sem"].set(str(counts["sem"]))
        self._kpi_vars["benef"].set(str(counts["benef"]))
        self._kpi_vars["near"].set(str(counts["near"]))
        self._kpi_vars["only_cp"].set(str(counts["only_cp"]))

        lbl_map = {
            "todos":   "Todos",
            "ok":      f"✅ Conciliado ({counts['ok']})",
            "sem":     f"Somente no DDA ({counts['sem']})",
            "benef":   f"Benef. diferentes ({counts['benef']})",
            "near":    f"Valores dif. ({counts['near']})",
            "only_cp": f"Somente no CP ({counts['only_cp']})",
        }
        for cod, lbl in lbl_map.items():
            self._filtro_btns[cod].config_text(lbl)

        self._aplicar_filtro("todos")

        alertas = counts["sem"] + counts["benef"] + counts["near"]
        self._status.set(
            f"✅ {counts['ok']} conciliados  |  "
            f"❌ {counts['sem']} somente no DDA  |  "
            f"⚠ {counts['benef']} beneficiários dif.  |  "
            f"≈ {counts['near']} valores dif.  |  "
            f"🔶 {counts['only_cp']} somente no CP  —  "
            f"{alertas} alerta(s) no total"
        )

    # ── Filtro ───────────────────────────────────────────────────────────────

    def _aplicar_filtro(self, filtro):
        self._filtro = filtro
        for cod, btn in self._filtro_btns.items():
            ativo = (cod == filtro)
            btn._cor  = CORES["accent"] if ativo else CORES["panel"]
            btn._ctxt = "#ffffff"       if ativo else CORES["text"]
            btn._desenhar(False)

        self._tree.delete(*self._tree.get_children())
        for vals, tag in self._all_rows:
            if filtro == "todos" or filtro == tag:
                self._tree.insert('', 'end', values=vals, tags=(tag,))

    # ── Exportar Excel ───────────────────────────────────────────────────────

    def _exportar(self):
        if not self._all_rows:
            messagebox.showwarning("Atenção", "Execute a comparação antes de exportar.")
            return

        path = filedialog.asksaveasfilename(
            title="Salvar relatório Excel",
            defaultextension=".xlsx",
            initialfile=f"COMPARATIVO_DDA_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Erro", "openpyxl não encontrado. Execute: pip install openpyxl")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo DDA"

        # Paleta de cores Excel
        fill_map = {t: PatternFill("solid", fgColor=c)
                    for t, c in self._XL_FILL.items()}
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin  = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            'Status', 'DDA — Beneficiário', 'DDA — CNPJ',
            'Nosso Número', 'Vencimento', 'DDA Valor',
            'CP — Fornecedor', 'CP — Duplicata', 'CP Valor', 'Observação',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border
        ws.row_dimensions[1].height = 32

        data_align = Alignment(vertical="center", wrap_text=False)
        wrap_align  = Alignment(vertical="center", wrap_text=True)

        for vals, tag in self._all_rows:
            ws.append(list(vals))
            row_num = ws.max_row
            fill = fill_map.get(tag)
            for i, cell in enumerate(ws[row_num], 1):
                if fill:
                    cell.fill = fill
                cell.border    = border
                cell.alignment = wrap_align if i == 1 else data_align

        # Larguras das colunas (caracteres)
        col_widths = [52, 36, 20, 22, 13, 14, 34, 16, 14, 30]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Congelar cabeçalho
        ws.freeze_panes = "A2"

        # Aba de resumo
        ws2 = wb.create_sheet("Resumo")
        counts = {t: sum(1 for _, tg in self._all_rows if tg == t)
                  for t in ("ok", "sem", "benef", "near", "only_cp")}
        total_dda = counts["ok"] + counts["sem"] + counts["benef"] + counts["near"]
        resumo_rows = [
            ("Situação", "Quantidade"),
            ("Total de boletos no DDA", total_dda),
            (self._ST["ok"],      counts["ok"]),
            (self._ST["sem"],     counts["sem"]),
            (self._ST["benef"],   counts["benef"]),
            (self._ST["near"],    counts["near"]),
            (self._ST["only_cp"], counts["only_cp"]),
        ]
        for i, row in enumerate(resumo_rows, 1):
            ws2.append(list(row))
            if i == 1:
                for cell in ws2[i]:
                    cell.font = Font(bold=True)
            else:
                tag_for_row = {
                    self._ST["ok"]:      "ok",
                    self._ST["sem"]:     "sem",
                    self._ST["benef"]:   "benef",
                    self._ST["near"]:    "near",
                    self._ST["only_cp"]: "only_cp",
                }.get(row[0])
                if tag_for_row:
                    for cell in ws2[i]:
                        cell.fill = fill_map[tag_for_row]
        ws2.column_dimensions["A"].width = 55
        ws2.column_dimensions["B"].width = 14

        ws2.append([])
        ws2.append(["Gerado em", datetime.now().strftime('%d/%m/%Y %H:%M')])

        wb.save(path)
        messagebox.showinfo("Exportado", f"Excel salvo em:\n{path}")
        self._status.set(f"Excel salvo: {os.path.basename(path)}")

    # ── Resetar ──────────────────────────────────────────────────────────────

    def _resetar(self):
        self._cp_path   = None
        self._dda_path  = None
        self._resultado = None
        self._all_rows  = []
        self._var_cp.set("Nenhum arquivo selecionado")
        self._var_dda.set("Nenhum arquivo selecionado")
        for k in self._kpi_vars:
            self._kpi_vars[k].set("—")
        self._tree.delete(*self._tree.get_children())
        lbl0 = {
            "todos":   "Todos",
            "ok":      "✅ Conciliado",
            "sem":     "Somente no DDA",
            "benef":   "Benef. diferentes",
            "near":    "Valores dif.",
            "only_cp": "Somente no CP",
        }
        for cod, lbl in lbl0.items():
            self._filtro_btns[cod].config_text(lbl)
        self._aplicar_filtro("todos")
        self._status.set("Selecione os arquivos e clique em Comparar.")




# ════════════════════════════════════════════════════════════════════════════════
#  MÓDULO 6 — CONCILIAÇÃO CARTÃO × DUPLICATAS EM ABERTO
# ════════════════════════════════════════════════════════════════════════════════

def _conc_parse_txt(path):
    """Lê o TXT de duplicatas em aberto do ERP e retorna lista de dicts."""
    dups = []
    pat = re.compile(
        r'(\d+)\s+([\d\/]+\.\d+)\s+\S*\s*\S*\s+'
        r'(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+'
        r'([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+R\$.*Doc:\s*(NF\s+\d+)'
    )
    with open(path, encoding='latin-1') as f:
        for line in f:
            m = pat.search(line)
            if m:
                dups.append({
                    'emp':        m.group(1).strip(),
                    'duplicata':  m.group(2).strip(),
                    'emissao':    m.group(3).strip(),
                    'vencimento': m.group(4).strip(),
                    'atraso':     m.group(5).strip(),
                    'valor':      float(m.group(6).replace('.', '').replace(',', '.')),
                    'valorMora':  float(m.group(7).replace('.', '').replace(',', '.')),
                    'valorTotal': float(m.group(8).replace('.', '').replace(',', '.')),
                    'nf':         m.group(9).strip(),
                })
    return dups


def _conc_parse_csv(path):
    """Lê o CSV de vendas no cartão e retorna lista de dicts (só aprovadas)."""
    rows = []
    for enc in ('latin-1', 'cp1252', 'utf-8-sig', 'utf-8'):
        try:
            with open(path, encoding=enc) as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    # Normaliza chaves para minúsculas
                    row_lower = {k.lower().strip(): v for k, v in row.items()}
                    status = row_lower.get('status da venda', '').lower().strip()
                    if status == 'aprovada':
                        raw = row_lower.get('valor da venda original', '0')
                        try:
                            val = float(str(raw).replace('.', '').replace(',', '.'))
                        except (ValueError, AttributeError):
                            continue
                        entry = dict(row_lower)
                        entry['_valor'] = val
                        rows.append(entry)
            break
        except Exception:
            continue
    return rows


def _conc_conciliar(dups, sales):
    """Cruza duplicatas com vendas. Retorna lista com campo 'status'."""
    results = []
    for dup in dups:
        data_dup = dup['emissao']
        valor    = dup['valor']
        # 1. Confirmado: data E valor batem
        exact = [s for s in sales
                 if s.get('data da venda', '') == data_dup
                 and abs(s['_valor'] - valor) < 0.01]
        if exact:
            results.append({**dup, 'status': 'confirmed', 'matches': exact})
            continue
        # 2. A verificar: só valor bate
        loose = [s for s in sales if abs(s['_valor'] - valor) < 0.01]
        if loose:
            results.append({**dup, 'status': 'possible', 'matches': loose})
            continue
        # 3. Não encontrado
        results.append({**dup, 'status': 'notfound', 'matches': []})
    return results


class FrameConciliacaoCartao(tk.Frame):
    """Modulo: Conciliacao Cartao x Duplicatas em Aberto.
    Layout em cards, espelhando o HTML Conciliacao_Cartao_Duplicatas.html.
    """

    # Paleta por status
    _PALETA = {
        "confirmed": {
            "badge_bg": "#dcfce7", "badge_fg": "#15803d",
            "val_fg":   "#15803d", "sec_fg": "#15803d",
            "label": "Venda encontrada na Rede",
            "sec":   "Venda encontrada na Rede — data e valor coincidem",
        },
        "possible": {
            "badge_bg": "#fef9c3", "badge_fg": "#92400e",
            "val_fg":   "#b45309", "sec_fg": "#b45309",
            "label": "A verificar",
            "sec":   "A verificar — mesmo valor, data diferente",
        },
        "notfound": {
            "badge_bg": "#fee2e2", "badge_fg": "#b91c1c",
            "val_fg":   "#b91c1c", "sec_fg": "#b91c1c",
            "label": "Nao consta vendas nesse filtro",
            "sec":   "Nao consta vendas nesse filtro",
        },
    }

    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg="#f4f5f7", **kwargs)
        self._csv_path = None
        self._txt_path = None
        self._results  = []
        self._build()

    # =========================================================================
    #  CONSTRUCAO DA INTERFACE
    # =========================================================================

    def _build(self):
        # --- Cabecalho azul-escuro (igual ao HTML) ---
        cab = tk.Frame(self, bg="#1F4E79")
        cab.pack(fill="x")
        tk.Label(cab,
                 text="  Conciliacao Cartao x Duplicatas em Aberto",
                 font=("Segoe UI", 13, "bold"), bg="#1F4E79", fg="#ffffff"
                 ).pack(side="left", padx=16, pady=(12, 4))
        tk.Label(cab,
                 text="Faca upload dos dois arquivos para gerar a analise automaticamente",
                 font=("Segoe UI", 8), bg="#1F4E79", fg="#aac4e8"
                 ).pack(side="left", anchor="s", pady=(0, 10))

        wrap = tk.Frame(self, bg="#f4f5f7")
        wrap.pack(fill="both", expand=True, padx=20, pady=12)

        # --- Selecao de arquivos ---
        frm_f = tk.Frame(wrap, bg="#ffffff",
                         highlightbackground="#dddddd", highlightthickness=1)
        frm_f.pack(fill="x", pady=(0, 10))

        self._var_csv = tk.StringVar(value="Nenhum arquivo selecionado")
        self._var_txt = tk.StringVar(value="Nenhum arquivo selecionado")
        self._file_row(frm_f, "Vendas no Cartao (.csv)",
                       "Relatorio da adquirente (Rede / Cielo / Stone ...)",
                       self._var_csv, self._sel_csv, 0)
        self._file_row(frm_f, "Duplicatas em Aberto (.txt)",
                       "Relatorio exportado do sistema ERP.",
                       self._var_txt, self._sel_txt, 1)

        # --- Botao principal ---
        frm_btns = tk.Frame(wrap, bg="#f4f5f7")
        frm_btns.pack(fill="x", pady=(0, 10))

        self._btn = BotaoRound(frm_btns, "Analisar conciliacao", self._executar,
                               largura=560, altura=40, raio=10,
                               cor="#1F4E79", cor_hover="#163c5e")
        self._btn.pack(side="left", expand=True)

        BotaoRound(frm_btns, "Limpar", self._resetar,
                   largura=110, altura=40, raio=10,
                   cor="#ffffff", cor_hover="#e5e7eb",
                   cor_texto="#666666"
                   ).pack(side="left", padx=(8, 0))

        # --- 4 cards de resumo ---
        self._frm_cards = tk.Frame(wrap, bg="#f4f5f7")
        self._frm_cards.pack(fill="x", pady=(0, 10))
        self._cval = {}
        self._csub = {}
        defs = [
            ("total",    "Total em aberto",               "#1F4E79", "duplicatas"),
            ("conf",     "Venda encontrada na Rede",       "#15803d", "confirmadas"),
            ("possible", "A verificar",                    "#b45309", "possiveis"),
            ("notfound", "Nao consta vendas nesse filtro", "#b91c1c", "sem correspondencia"),
        ]
        for key, titulo, cor, sub0 in defs:
            c = tk.Frame(self._frm_cards, bg="#ffffff",
                         highlightbackground="#e0e0e0", highlightthickness=1)
            c.pack(side="left", expand=True, fill="x", padx=(0, 10), pady=2)
            tk.Label(c, text=titulo, font=("Segoe UI", 8),
                     bg="#ffffff", fg="#888888").pack(padx=12, pady=(10, 2), anchor="w")
            vl = tk.Label(c, text="R$ 0,00",
                          font=("Segoe UI", 17, "bold"), bg="#ffffff", fg=cor)
            vl.pack(padx=12, pady=(0, 2), anchor="w")
            sl = tk.Label(c, text=sub0, font=("Segoe UI", 7),
                          bg="#ffffff", fg="#bbbbbb")
            sl.pack(padx=12, pady=(0, 10), anchor="w")
            self._cval[key] = vl
            self._csub[key] = sl

        # --- Area scrollavel de resultados ---
        self._res_outer = tk.Frame(wrap, bg="#f4f5f7")
        # Comeca oculta; aparece apos analise

        self._canvas = tk.Canvas(self._res_outer, bg="#f4f5f7",
                                  highlightthickness=0, bd=0)
        self._sb = ttk.Scrollbar(self._res_outer, orient="vertical",
                                  command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._sb.set)
        self._sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._inner = tk.Frame(self._canvas, bg="#f4f5f7")
        self._win_id = self._canvas.create_window((0, 0), window=self._inner, anchor="nw")

        self._inner.bind("<Configure>",  self._on_inner_cfg)
        self._canvas.bind("<Configure>", self._on_canvas_cfg)

        # --- Nota de rodape ---
        self._nota_var = tk.StringVar()
        self._nota_lbl = tk.Label(wrap, textvariable=self._nota_var,
                                   font=("Segoe UI", 7), bg="#f4f5f7", fg="#aaaaaa",
                                   wraplength=800, justify="left", anchor="w")
        # aparece apos analise

        # --- Status ---
        self._status = tk.StringVar(value="Selecione os arquivos e clique em Analisar conciliacao.")
        tk.Label(self, textvariable=self._status,
                 font=("Segoe UI", 8), bg="#f4f5f7", fg="#888888", anchor="w"
                 ).pack(fill="x", padx=22, pady=(0, 6))

    # =========================================================================
    #  HELPERS DE LAYOUT
    # =========================================================================

    def _on_inner_cfg(self, event):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_cfg(self, event):
        self._canvas.itemconfig(self._win_id, width=event.width)

    def _scroll(self, event):
        self._canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _file_row(self, parent, titulo, dica, var, cmd, row):
        frm = tk.Frame(parent, bg="#ffffff")
        frm.pack(fill="x", padx=14, pady=(8 if row == 0 else 0, 8 if row == 1 else 0))

        ico = ">" if row == 0 else ">"
        tk.Label(frm, text=titulo, font=("Segoe UI", 9, "bold"),
                 bg="#ffffff", fg="#333333", anchor="w"
                 ).pack(side="left")
        tk.Label(frm, text="  —  " + dica, font=("Segoe UI", 8),
                 bg="#ffffff", fg="#aaaaaa"
                 ).pack(side="left")
        tk.Label(frm, textvariable=var, font=("Segoe UI", 8),
                 fg="#2563eb", bg="#ffffff", anchor="e"
                 ).pack(side="right", padx=(0, 8))
        BotaoRound(frm, "Selecionar...", cmd,
                   largura=100, altura=26, raio=6,
                   cor="#f3f4f6", cor_hover="#1F4E79",
                   cor_texto="#333333", fonte=("Segoe UI", 8)
                   ).pack(side="right")

        if row == 0:
            tk.Frame(parent, bg="#f0f0f0", height=1).pack(fill="x", padx=14)

    def _fmt(self, v):
        return "R$ " + "{:,.2f}".format(v).replace(",", "X").replace(".", ",").replace("X", ".")

    # =========================================================================
    #  SELECAO DE ARQUIVOS
    # =========================================================================

    def _sel_csv(self):
        f = filedialog.askopenfilename(title="Vendas no Cartao",
                                       filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if f:
            self._csv_path = f
            self._var_csv.set(os.path.basename(f))

    def _sel_txt(self):
        f = filedialog.askopenfilename(title="Duplicatas em Aberto",
                                       filetypes=[("Texto", "*.txt"), ("Todos", "*.*")])
        if f:
            self._txt_path = f
            self._var_txt.set(os.path.basename(f))

    # =========================================================================
    #  EXECUCAO
    # =========================================================================

    def _executar(self):
        if not self._csv_path or not self._txt_path:
            messagebox.showwarning("Atencao",
                                   "Selecione os dois arquivos antes de analisar.")
            return
        self._btn.config_text("Analisando...")
        self.update()
        try:
            dups  = _conc_parse_txt(self._txt_path)
            sales = _conc_parse_csv(self._csv_path)

            if not dups:
                raise ValueError(
                    "Nao foi possivel ler duplicatas do .txt.\n"
                    "Verifique se o arquivo contem o campo 'Doc: NF XXXXX'.")
            if not sales:
                raise ValueError(
                    "Nao foi possivel ler vendas do .csv\n"
                    "(nenhuma venda 'Aprovada' encontrada).\n"
                    "Verifique o formato e o separador (;).")

            self._results = _conc_conciliar(dups, sales)
            conf = [r for r in self._results if r["status"] == "confirmed"]
            poss = [r for r in self._results if r["status"] == "possible"]
            nf   = [r for r in self._results if r["status"] == "notfound"]

            total_v = sum(r["valor"] for r in self._results)
            conf_v  = sum(r["valor"] for r in conf)
            poss_v  = sum(r["valor"] for r in poss)
            nf_v    = sum(r["valor"] for r in nf)

            n = len(self._results)
            self._cval["total"].config(text=self._fmt(total_v))
            self._csub["total"].config(text="{} duplicata{}".format(n, "s" if n != 1 else ""))
            self._cval["conf"].config(text=self._fmt(conf_v))
            self._csub["conf"].config(text="{} confirmada{}".format(
                len(conf), "s" if len(conf) != 1 else ""))
            self._cval["possible"].config(text=self._fmt(poss_v))
            self._csub["possible"].config(text="{} possivel{}".format(
                len(poss), "is" if len(poss) != 1 else ""))
            self._cval["notfound"].config(text=self._fmt(nf_v))
            self._csub["notfound"].config(text="{} sem correspondencia".format(len(nf)))

            self._renderizar(conf, poss, nf)

            self._status.set(
                "Total: {}  |  {} confirmada(s)  |  {} a verificar  |  {} nao encontrada(s)".format(
                    self._fmt(total_v), len(conf), len(poss), len(nf)))

        except Exception as e:
            messagebox.showerror("Erro na analise", str(e))
            self._status.set("Erro: {}".format(e))
        finally:
            self._btn.config_text("Analisar conciliacao")

    # =========================================================================
    #  RENDERIZACAO DOS RESULTADOS
    # =========================================================================

    def _renderizar(self, conf, poss, nf):
        # Limpa area anterior
        for w in self._inner.winfo_children():
            w.destroy()
        self._canvas.unbind_all("<MouseWheel>")

        secoes = [
            (conf, "confirmed"),
            (poss, "possible"),
            (nf,   "notfound"),
        ]
        for items, status in secoes:
            if not items:
                continue
            pal = self._PALETA[status]

            # -- Titulo de secao (borda colorida a esquerda) --
            sep_outer = tk.Frame(self._inner, bg="#f4f5f7")
            sep_outer.pack(fill="x", pady=(14, 6))
            tk.Frame(sep_outer, bg=pal["sec_fg"], width=3).pack(side="left", fill="y")
            tk.Label(sep_outer,
                     text="{} ({})".format(pal["sec"], len(items)),
                     font=("Segoe UI", 8, "bold"),
                     bg="#f4f5f7", fg=pal["sec_fg"]
                     ).pack(side="left", padx=8, pady=3)

            for r in items:
                self._card(r, status, pal)

        # Nota de rodape
        self._nota_var.set(
            '"Nao consta vendas nesse filtro" pode indicar venda realizada em outra '
            'adquirente (Cielo, Stone, GetNet) ou baixa manual ja efetuada fora do sistema Rede.')
        self._nota_lbl.pack(fill="x", pady=(8, 4))

        # Mostra area de resultados e habilita scroll
        self._res_outer.pack(fill="both", expand=True, pady=(0, 4))
        self._canvas.bind_all("<MouseWheel>", self._scroll)
        self.bind("<Destroy>", lambda e: self._canvas.unbind_all("<MouseWheel>"))
        self._canvas.yview_moveto(0)

    def _card(self, r, status, pal):
        m = r["matches"][0] if r["matches"] else {}

        card = tk.Frame(self._inner, bg="#ffffff",
                        highlightbackground="#e0e0e0", highlightthickness=1)
        card.pack(fill="x", pady=(0, 10))

        # ---- Cabecalho do card ----
        head = tk.Frame(card, bg="#ffffff")
        head.pack(fill="x", padx=14, pady=(10, 8))

        lft = tk.Frame(head, bg="#ffffff")
        lft.pack(side="left")

        # Badge colorido
        tk.Label(lft, text=pal["label"],
                 font=("Segoe UI", 8, "bold"),
                 bg=pal["badge_bg"], fg=pal["badge_fg"],
                 padx=8, pady=2
                 ).pack(side="left")
        tk.Label(lft, text="  " + r["nf"],
                 font=("Segoe UI", 11, "bold"),
                 bg="#ffffff", fg="#1F4E79"
                 ).pack(side="left", padx=(8, 0))
        tk.Label(lft, text="  · duplicata " + r["duplicata"],
                 font=("Segoe UI", 9),
                 bg="#ffffff", fg="#888888"
                 ).pack(side="left")

        # Valor a direita
        tk.Label(head, text=self._fmt(r["valor"]),
                 font=("Segoe UI", 13, "bold"),
                 bg="#ffffff", fg=pal["val_fg"]
                 ).pack(side="right")

        # Linha separadora
        tk.Frame(card, bg="#f0f0f0", height=1).pack(fill="x")

        # ---- Corpo: dois paineis lado a lado ----
        body = tk.Frame(card, bg="#ffffff")
        body.pack(fill="x")
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        # -- Painel esquerdo: sistema --
        pl = tk.Frame(body, bg="#ffffff")
        pl.grid(row=0, column=0, sticky="nsew", padx=(14, 6), pady=12)

        tk.Label(pl, text="DADOS DO SISTEMA (DUPLICATA)",
                 font=("Segoe UI", 7, "bold"),
                 bg="#ffffff", fg="#aaaaaa"
                 ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        atraso_int = int(r["atraso"]) if str(r["atraso"]).isdigit() else 0
        sys_rows = [
            ("Documento",       r["nf"],              "#1F4E79"),
            ("Duplicata",       r["duplicata"],        "#222222"),
            ("Empresa",         r["emp"],              "#222222"),
            ("Data emissao",    r["emissao"],          "#222222"),
            ("Vencimento",      r["vencimento"],       "#222222"),
            ("Atraso",
             "{} dia{}".format(r["atraso"], "s" if atraso_int != 1 else ""),
             "#b45309" if atraso_int > 0 else "#222222"),
            ("Valor em aberto", self._fmt(r["valor"]),     "#222222"),
            ("Valor mora",      self._fmt(r["valorMora"]), "#222222"),
            ("Total com mora",  self._fmt(r["valorTotal"]), "#222222"),
        ]
        for i, (lbl, val, fg) in enumerate(sys_rows, start=1):
            tk.Label(pl, text=lbl + ":",
                     font=("Segoe UI", 8), bg="#ffffff", fg="#999999", anchor="w"
                     ).grid(row=i, column=0, sticky="w", padx=(0, 12))
            tk.Label(pl, text=val,
                     font=("Segoe UI", 8, "bold"), bg="#ffffff", fg=fg, anchor="w"
                     ).grid(row=i, column=1, sticky="w")

        # Divisor vertical
        tk.Frame(body, bg="#f0f0f0", width=1).grid(row=0, column=0, sticky="nse",
                                                    padx=(0, 0), pady=12)

        # -- Painel direito: Rede --
        pr = tk.Frame(body, bg="#ffffff")
        pr.grid(row=0, column=1, sticky="nsew", padx=(14, 14), pady=12)

        tk.Label(pr, text="DADOS DA VENDA — REDE",
                 font=("Segoe UI", 7, "bold"),
                 bg="#ffffff", fg="#aaaaaa"
                 ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        if m:
            def _ddiff(d1, d2):
                try:
                    from datetime import datetime as _dt
                    _p = lambda s: _dt.strptime(s, "%d/%m/%Y")
                    return (_p(d2) - _p(d1)).days
                except Exception:
                    return None

            dv    = m.get("data da venda", "")
            diff  = _ddiff(r["emissao"], dv)
            if diff is None:
                diff_txt = "—"
                diff_ok  = True
            elif diff == 0:
                diff_txt = "—"
                diff_ok  = True
            else:
                diff_txt = "{} dia{} {}".format(
                    abs(diff), "s" if abs(diff) != 1 else "",
                    "apos" if diff > 0 else "antes")
                diff_ok  = False

            parc_raw  = m.get("numero de parcelas", m.get(
                               "numero de parcelas", "1")).strip()
            parc_txt  = "a vista" if parc_raw in ("", "0", "1") else parc_raw + "x"
            modalidade = "{} {} · {}".format(
                m.get("modalidade", ""), m.get("tipo", ""), parc_txt).strip()
            vliq = m.get("valor liquido", m.get("valor liquido", ""))

            def _kv(d, *keys):
                for k in keys:
                    v = d.get(k, "")
                    if v:
                        return v
                return "—"

            rede_rows = [
                ("Data da venda",  dv or "—",
                 "#15803d" if (diff is not None and diff == 0) else "#b45309"),
                ("Hora",           _kv(m, "hora da venda"),   "#222222"),
                ("Dif. de data",   diff_txt,
                 "#15803d" if diff_ok else "#b45309"),
                ("Modalidade",     modalidade,                 "#222222"),
                ("Bandeira",       _kv(m, "bandeira"),         "#222222"),
                ("Cartao",         _kv(m, "numero do cartao",
                                       "numero do cartao"),    "#222222"),
                ("NSU / CV",       _kv(m, "nsu/cv"),           "#1F4E79"),
                ("Autorizacao",    _kv(m, "numero da autorizacao",
                                       "numero da autorizacao (auto)"), "#222222"),
                ("Valor bruto",    self._fmt(m["_valor"]),     "#222222"),
                ("Valor liquido",  vliq if vliq else "—", "#222222"),
            ]
            for i, (lbl, val, fg) in enumerate(rede_rows, start=1):
                tk.Label(pr, text=lbl + ":",
                         font=("Segoe UI", 8), bg="#ffffff", fg="#999999", anchor="w"
                         ).grid(row=i, column=0, sticky="w", padx=(0, 12))
                tk.Label(pr, text=val,
                         font=("Segoe UI", 8, "bold"), bg="#ffffff", fg=fg, anchor="w"
                         ).grid(row=i, column=1, sticky="w")
        else:
            tk.Label(pr,
                     text="Nenhuma transacao com o valor de {}\nencontrada no periodo.".format(
                         self._fmt(r["valor"])),
                     font=("Segoe UI", 8),
                     bg="#ffffff", fg="#bbbbbb",
                     justify="left"
                     ).grid(row=1, column=0, columnspan=2, sticky="w", pady=8)

    # =========================================================================
    #  RESET
    # =========================================================================

    def _resetar(self):
        self._csv_path = None
        self._txt_path = None
        self._results  = []
        self._var_csv.set("Nenhum arquivo selecionado")
        self._var_txt.set("Nenhum arquivo selecionado")
        orig = {
            "total": ("R$ 0,00", "duplicatas"),
            "conf":  ("R$ 0,00", "confirmadas"),
            "possible": ("R$ 0,00", "possiveis"),
            "notfound": ("R$ 0,00", "sem correspondencia"),
        }
        for key, (v, s) in orig.items():
            self._cval[key].config(text=v)
            self._csub[key].config(text=s)
        for w in self._inner.winfo_children():
            w.destroy()
        self._canvas.unbind_all("<MouseWheel>")
        self._res_outer.pack_forget()
        self._nota_lbl.pack_forget()
        self._status.set("Selecione os arquivos e clique em Analisar conciliacao.")


# ================================================================================
#  APLICACAO PRINCIPAL — NAVEGACAO LATERAL
# ================================================================================

MENU = [
    ("\U0001f4ca", "Conferir Duplicatas", FrameConferirDuplicatas),
    ("\U0001f6e1",  "Seguro de Vida",      FrameSeguroVida),
    ("\U0001f4bc", "Contas a Pagar",      FrameContasPagar),
    ("\U0001f50d", "Creditos em Aberto",  FrameVerificadorCreditos),
    ("\U0001f3e6", "Comparador DDA",      FrameComparadorDDA),
    ("\U0001f4b3", "Conciliacao Cartao",  FrameConciliacaoCartao),
]


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Ferramentas Krambeck")
        self.configure(bg=CORES["bg"])
        self.resizable(True, True)
        self.minsize(960, 640)
        self._center(1060, 700)
        self._frames     = {}
        self._nav_btns   = {}
        self._indicators = {}
        self._rows       = {}
        self._build()
        self._show(MENU[0][2])

    def _center(self, w, h):
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _carregar_logo(self):
        self._logo_header_kr  = None
        self._logo_header_an  = None
        self._logo_sidebar_kr = None
        self._logo_sidebar_an = None
        if not _PIL_OK:
            return

        base = os.path.dirname(os.path.abspath(__file__))
        arqs = {
            "kr": os.path.join(base, "LOGO KRAMBECK 2.png"),
            "an": os.path.join(base, "LOGO ANCORA.png"),
        }

        SB_BG = (30, 41, 59, 255)
        HD_BG = (255, 255, 255, 255)

        def _comp(img, bg_rgba, target_h, alpha_pct=1.0):
            w = int(img.width * (target_h / img.height))
            img = img.resize((w, target_h), Image.LANCZOS)
            if alpha_pct < 1.0:
                r, g, b, a = img.split()
                a = a.point(lambda v: int(v * alpha_pct))
                img.putalpha(a)
            bg = Image.new("RGBA", img.size, bg_rgba)
            return Image.alpha_composite(bg, img)

        for key, path in arqs.items():
            if not os.path.exists(path):
                continue
            try:
                orig = Image.open(path).convert("RGBA")
                h_img = _comp(orig, HD_BG, 34)
                s_img = _comp(orig, SB_BG, 28, alpha_pct=0.75)
                if key == "kr":
                    self._logo_header_kr  = ImageTk.PhotoImage(h_img)
                    self._logo_sidebar_kr = ImageTk.PhotoImage(s_img)
                else:
                    self._logo_header_an  = ImageTk.PhotoImage(h_img)
                    self._logo_sidebar_an = ImageTk.PhotoImage(s_img)
            except Exception:
                pass

    def _build(self):
        self._carregar_logo()

        header = tk.Frame(self, bg=CORES["header"], height=58)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="  Ferramentas Krambeck",
                 font=("Segoe UI", 15, "bold"),
                 fg=CORES["text"], bg=CORES["header"]
                 ).pack(side="left", padx=16, pady=10)
        logo_frame_h = tk.Frame(header, bg=CORES["header"])
        logo_frame_h.pack(side="right", padx=14, pady=10)
        if self._logo_header_an:
            tk.Label(logo_frame_h, image=self._logo_header_an,
                     bg=CORES["header"], bd=0).pack(side="right", padx=(6, 0))
        if self._logo_header_kr:
            tk.Label(logo_frame_h, image=self._logo_header_kr,
                     bg=CORES["header"], bd=0).pack(side="right", padx=(0, 6))

        tk.Frame(self, bg=CORES["accent"], height=2).pack(fill="x")

        body = tk.Frame(self, bg=CORES["bg"])
        body.pack(fill="both", expand=True)

        sidebar = tk.Frame(body, bg=CORES["sidebar"], width=214)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text="MENU", font=("Segoe UI", 8, "bold"),
                 fg=CORES["muted"], bg=CORES["sidebar"]
                 ).pack(padx=16, pady=(20, 6), anchor="w")

        for icon, label, cls in MENU:
            row_frame = tk.Frame(sidebar, bg=CORES["sidebar"], height=46)
            row_frame.pack(fill="x", pady=1)
            row_frame.pack_propagate(False)
            self._rows[cls] = row_frame

            ind = tk.Frame(row_frame, bg=CORES["sidebar"], width=4)
            ind.pack(side="left", fill="y")
            self._indicators[cls] = ind

            ico_lbl = tk.Label(
                row_frame, text=icon, font=("Segoe UI Emoji", 13),
                fg=CORES["muted"], bg=CORES["sidebar"],
                width=2, anchor="center")
            ico_lbl.pack(side="left", padx=(6, 0))

            btn = tk.Button(
                row_frame, text=label, font=("Segoe UI", 10),
                fg=CORES["muted"], bg=CORES["sidebar"],
                relief="flat", bd=0, padx=8, pady=0,
                anchor="w", cursor="hand2",
                activebackground=CORES["panel"],
                activeforeground=CORES["text"],
                command=lambda c=cls: self._show(c))
            btn.pack(side="left", fill="both", expand=True)
            self._nav_btns[cls] = (btn, ico_lbl)

        logo_frame_s = tk.Frame(sidebar, bg=CORES["sidebar"])
        logo_frame_s.pack(side="bottom", pady=16)
        if self._logo_sidebar_kr:
            tk.Label(logo_frame_s, image=self._logo_sidebar_kr,
                     bg=CORES["sidebar"], bd=0).pack(side="left", padx=4)
        if self._logo_sidebar_an:
            tk.Label(logo_frame_s, image=self._logo_sidebar_an,
                     bg=CORES["sidebar"], bd=0).pack(side="left", padx=4)

        content = tk.Frame(body, bg=CORES["bg"])
        content.pack(side="left", fill="both", expand=True)
        content.rowconfigure(0, weight=1)
        content.columnconfigure(0, weight=1)

        for _, __, cls in MENU:
            frm = cls(content)
            frm.grid(row=0, column=0, sticky="nsew")
            self._frames[cls] = frm

    def _show(self, cls):
        _ACTIVE_ROW = "#2d4a7a"
        _ACTIVE_TXT = "#ffffff"
        _ACTIVE_ICO = "#93c5fd"
        for c, (btn, ico) in self._nav_btns.items():
            row = self._rows[c]
            if c is cls:
                row.config(bg=_ACTIVE_ROW)
                btn.config(bg=_ACTIVE_ROW, fg=_ACTIVE_TXT,
                           font=("Segoe UI", 10, "bold"),
                           activebackground=_ACTIVE_ROW,
                           activeforeground=_ACTIVE_TXT)
                ico.config(bg=_ACTIVE_ROW, fg=_ACTIVE_ICO)
                self._indicators[c].config(bg=CORES["accent"])
            else:
                row.config(bg=CORES["sidebar"])
                btn.config(bg=CORES["sidebar"], fg=CORES["muted"],
                           font=("Segoe UI", 10),
                           activebackground=CORES["sidebar"],
                           activeforeground="#ffffff")
                ico.config(bg=CORES["sidebar"], fg=CORES["muted"])
                self._indicators[c].config(bg=CORES["sidebar"])
        self._frames[cls].tkraise()


if __name__ == "__main__":
    app = App()
    app.mainloop()
